import re
from django.shortcuts import render, redirect, get_object_or_404
from django.http import StreamingHttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy, reverse
from django.contrib.auth import login
from django.views import View
from django.db import transaction
from decimal import Decimal
import json
from .models import *
from .forms import SignUpForm, BrandForm, ProductGroupForm, SupplierForm, ProductForm, CustomerForm, InvoiceForm, InvoiceDetailFormSet
from shared.mixins import StaffRequiredMixin, PermissionRequiredMixin
from django.utils.decorators import method_decorator
from shared.decorators import audit_action, permission_required

import threading
from reportlab.pdfgen import canvas
from reportlab.lib import colors

_local_data = threading.local()

def set_current_report_title(title):
    _local_data.current_title = title

def get_current_report_title():
    return getattr(_local_data, 'current_title', "Sistema de Ventas - Reporte")

class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor("#7F7F7F"))
        self.setStrokeColor(colors.HexColor("#D9D9D9"))
        self.setLineWidth(0.5)
        self.line(36, 756, 576, 756)
        
        # Obtener título dinámico
        title = get_current_report_title()
        self.drawString(36, 762, title)
        
        self.line(36, 54, 576, 54)
        page_text = f"Página {self._pageNumber} de {page_count}"
        self.drawRightString(576, 42, page_text)
        self.drawString(36, 42, "Reporte generado automáticamente")
        self.restoreState()


# === REGISTRO ===
class SignUpView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    form_class = SignUpForm
    template_name = 'registration/signup.html'
    success_url = reverse_lazy('billing:brand_list')

    def test_func(self):
        # Desactivado para el público general; solo superusuarios (o admin) pueden acceder.
        return self.request.user.is_superuser

    def form_valid(self, form):
        response = super().form_valid(form)
        login(self.request, self.object)
        return response

@login_required
def home(request):
    """Vista principal del sistema. Muestra resumen personalizado según el rol."""
    from django.db.models import Sum
    from django.contrib.auth.models import User
    from purchasing.models import Purchase
    from pagos.models import CobroFactura, PagoCompra

    user = request.user
    
    # Determinar los grupos del usuario
    groups = list(user.groups.values_list('name', flat=True))
    is_admin = user.is_superuser or 'Administrador' in groups
    is_gerente = 'Gerente' in groups
    is_ventas = 'Ventas' in groups
    is_compras = 'Compras' in groups

    context = {
        'groups': groups,
        'is_admin': is_admin,
        'is_gerente': is_gerente,
        'is_ventas': is_ventas,
        'is_compras': is_compras,
    }

    # Cargar datos compartidos o específicos según el rol
    if is_admin:
        context.update({
            'total_users': User.objects.count(),
            'recent_users': User.objects.order_by('-date_joined')[:5],
            'recent_logins': User.objects.filter(last_login__isnull=False).order_by('-last_login')[:5],
            'low_stock_count': Product.objects.filter(stock__lte=5, is_active=True).count(),
            'total_invoices': Invoice.objects.filter(is_active=True).count(),
            'total_purchases': Purchase.objects.filter(is_active=True).count(),
        })

    if is_gerente or is_admin:
        # Ventas totales
        invoices = Invoice.objects.filter(is_active=True)
        total_ventas = invoices.aggregate(Sum('total'))['total__sum'] or Decimal('0.00')
        # Compras totales
        purchases = Purchase.objects.filter(is_active=True)
        total_compras = purchases.aggregate(Sum('total'))['total__sum'] or Decimal('0.00')
        
        # Saldos pendientes
        saldo_por_cobrar = invoices.filter(tipo_pago='CREDITO', estado='PENDIENTE').aggregate(Sum('saldo'))['saldo__sum'] or Decimal('0.00')
        saldo_por_pagar = purchases.filter(tipo_pago='CREDITO', estado='PENDIENTE').aggregate(Sum('saldo'))['saldo__sum'] or Decimal('0.00')

        # Cobros y Pagos realizados
        total_cobrado = CobroFactura.objects.aggregate(Sum('valor'))['valor__sum'] or Decimal('0.00')
        total_pagado = PagoCompra.objects.aggregate(Sum('valor'))['valor__sum'] or Decimal('0.00')

        context.update({
            'total_ventas': total_ventas,
            'total_compras': total_compras,
            'saldo_por_cobrar': saldo_por_cobrar,
            'saldo_por_pagar': saldo_por_pagar,
            'total_cobrado': total_cobrado,
            'total_pagado': total_pagado,
            'total_products': Product.objects.filter(is_active=True).count(),
            'total_customers': Customer.objects.filter(is_active=True).count(),
            'total_suppliers': Supplier.objects.filter(is_active=True).count(),
        })

    if is_ventas or is_admin or is_gerente:
        context.update({
            'total_customers': Customer.objects.filter(is_active=True).count(),
            'total_invoices_sales': Invoice.objects.filter(is_active=True).count(),
            'recent_invoices': Invoice.objects.filter(is_active=True).select_related('customer').order_by('-invoice_date')[:5],
            'pending_cobros': Invoice.objects.filter(tipo_pago='CREDITO', estado='PENDIENTE', is_active=True).select_related('customer').order_by('-invoice_date')[:5],
        })

    if is_compras or is_admin or is_gerente:
        context.update({
            'low_stock': Product.objects.filter(stock__lte=5, is_active=True).select_related('brand'),
            'total_suppliers': Supplier.objects.filter(is_active=True).count(),
            'recent_purchases': Purchase.objects.filter(is_active=True).select_related('supplier').order_by('-purchase_date')[:5],
            'pending_pagos': Purchase.objects.filter(tipo_pago='CREDITO', estado='PENDIENTE', is_active=True).select_related('supplier').order_by('-purchase_date')[:5],
        })

    return render(request, 'billing/home.html', context)


# === BRAND (FBV) ===
@login_required
@permission_required('billing.view_brand')
@audit_action('LIST_BRANDS')  
def brand_list(request):
    search_field = request.GET.get('search_field', 'all').strip()
    search_value = request.GET.get('search_value', '').strip()
    
    brands = Brand.objects.all()
    
    if search_value:
        from django.db.models import Q
        if search_field == 'name':
            brands = brands.filter(name__icontains=search_value)
        elif search_field == 'description':
            brands = brands.filter(description__icontains=search_value)
        elif search_field == 'active':
            val = search_value.lower()
            if val in ['activo', 'activa', 'si', 'sí', '1', 'true']:
                brands = brands.filter(is_active=True)
            elif val in ['inactivo', 'inactiva', 'no', '0', 'false']:
                brands = brands.filter(is_active=False)
        else: # 'all'
            q_filters = Q(name__icontains=search_value) | Q(description__icontains=search_value)
            val = search_value.lower()
            if val in ['activo', 'activa', 'si', 'sí', '1', 'true']:
                q_filters |= Q(is_active=True)
            elif val in ['inactivo', 'inactiva', 'no', '0', 'false']:
                q_filters |= Q(is_active=False)
            brands = brands.filter(q_filters)
            
    # Paginación de hasta 10 marcas
    from django.core.paginator import Paginator
    paginator = Paginator(brands, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'billing/brand_list.html', {
        'brands': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'is_paginated': page_obj.has_other_pages(),
        'search_field': search_field,
        'search_value': search_value
    })

@login_required
@permission_required('billing.add_brand')
@audit_action('CREATE_BRAND')  
def brand_create(request):
    if request.method == 'POST':
        form = BrandForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Marca creada con éxito!')
            return redirect('billing:brand_list')
    else: form = BrandForm()
    return render(request, 'billing/brand_form.html', {'form':form, 'title':'Crear Marca'})

@login_required
@permission_required('billing.change_brand')
@audit_action('UPDATE_BRAND')  
def brand_update(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    if request.method == 'POST':
        form = BrandForm(request.POST, instance=brand)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Marca actualizada con éxito!')
            return redirect('billing:brand_list')
    else: form = BrandForm(instance=brand)
    return render(request, 'billing/brand_form.html', {'form':form, 'title':'Editar Marca'})

@login_required
@permission_required('billing.delete_brand')
@audit_action('DELETE_BRAND')  
def brand_delete(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    if request.method == 'POST':
        brand.delete()
        messages.success(request, '¡Marca eliminada con éxito!')
        return redirect('billing:brand_list')
    return render(request, 'billing/brand_confirm_delete.html', {'object': brand})

# =============================================
# CRUD DE INVOICE - VISTAS BASADAS EN FUNCIONES
# (Requiere FBV porque usa formsets complejos)
# =============================================

@login_required
@permission_required('billing.view_invoice')
def invoice_list(request):
    """Lista todas las facturas con sus totales."""
    invoices = Invoice.objects.select_related('customer').all()
    return render(request, 'billing/invoice_list.html', {'items': invoices})


@login_required
@permission_required('billing.add_invoice')
def invoice_create(request):
    """Crea factura con sus líneas de detalle."""
    if request.method == 'POST':
        form = InvoiceForm(request.POST)
        formset = InvoiceDetailFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            # Guardar factura (sin commit para asignar totales)
            invoice = form.save(commit=False)
            invoice.save()

            # Asignar la factura al formset y guardar detalles
            formset.instance = invoice
            details = formset.save()

            # Calcular totales
            subtotal = sum(d.subtotal for d in invoice.details.all())
            invoice.subtotal = subtotal
            invoice.tax = subtotal * Decimal('0.15')  # IVA 15%
            invoice.total = invoice.subtotal + invoice.tax
            invoice.save()

            messages.success(request, f'¡Factura #{invoice.id} creada con éxito! Total: ${invoice.total}')
            return redirect('billing:invoice_list')
    else:
        form = InvoiceForm()
        formset = InvoiceDetailFormSet()

    return render(request, 'billing/invoice_form.html', {
        'form': form,
        'formset': formset,
        'title': 'Crear Factura',
    })


@login_required
@permission_required('billing.view_invoice')
def invoice_detail(request, pk):
    """Muestra el detalle completo de una factura."""
    invoice = get_object_or_404(
        Invoice.objects.select_related('customer')
                       .prefetch_related('details__product'),
        pk=pk
    )
    return render(request, 'billing/invoice_detail.html', {'invoice': invoice})


@login_required
@permission_required('billing.delete_invoice')
def invoice_delete(request, pk):
    """Elimina una factura y todos sus detalles (CASCADE)."""
    invoice = get_object_or_404(Invoice, pk=pk)
    if request.method == 'POST':
        invoice_id = invoice.id
        invoice.delete()
        messages.success(request, f'¡Factura #{invoice_id} eliminada con éxito!')
        return redirect('billing:invoice_list')
    return render(request, 'billing/invoice_confirm_delete.html', {'object': invoice})


# === PRODUCTGROUP (CBV) ===
class ProductGroupListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'billing.view_productgroup'
    model = ProductGroup
    template_name = 'billing/productgroup_list.html'
    context_object_name = 'items'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()
        search_field = self.request.GET.get('search_field', 'all').strip()
        search_value = self.request.GET.get('search_value', '').strip()
        if search_value:
            from django.db.models import Q
            if search_field == 'name':
                queryset = queryset.filter(name__icontains=search_value)
            elif search_field == 'active':
                val = search_value.lower()
                if val in ['activo', 'activa', 'si', 'sí', '1', 'true']:
                    queryset = queryset.filter(is_active=True)
                elif val in ['inactivo', 'inactiva', 'no', '0', 'false']:
                    queryset = queryset.filter(is_active=False)
            else: # 'all'
                q_filters = Q(name__icontains=search_value)
                val = search_value.lower()
                if val in ['activo', 'activa', 'si', 'sí', '1', 'true']:
                    q_filters |= Q(is_active=True)
                elif val in ['inactivo', 'inactiva', 'no', '0', 'false']:
                    q_filters |= Q(is_active=False)
                queryset = queryset.filter(q_filters)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_field'] = self.request.GET.get('search_field', 'all').strip()
        context['search_value'] = self.request.GET.get('search_value', '').strip()
        return context

class ProductGroupCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'billing.add_productgroup'
    model = ProductGroup; form_class = ProductGroupForm; template_name = 'billing/productgroup_form.html'; success_url = reverse_lazy('billing:productgroup_list')
class ProductGroupUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'billing.change_productgroup'
    model = ProductGroup; form_class = ProductGroupForm; template_name = 'billing/productgroup_form.html'; success_url = reverse_lazy('billing:productgroup_list')
class ProductGroupDeleteView(LoginRequiredMixin, PermissionRequiredMixin, StaffRequiredMixin, DeleteView):
    permission_required = 'billing.delete_productgroup'
    model = ProductGroup
    template_name = 'billing/productgroup_confirm_delete.html'
    success_url = reverse_lazy('billing:productgroup_list')
    staff_redirect_url = '/groups/'  # Redirige aquí si no es staff


# === SUPPLIER (CBV) ===
class SupplierListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'billing.view_supplier'
    model = Supplier
    template_name = 'billing/supplier_list.html'
    context_object_name = 'items'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()
        search_field = self.request.GET.get('search_field', 'all').strip()
        search_value = self.request.GET.get('search_value', '').strip()
        if search_value:
            from django.db.models import Q
            if search_field == 'name':
                queryset = queryset.filter(name__icontains=search_value)
            elif search_field == 'contact_name':
                queryset = queryset.filter(contact_name__icontains=search_value)
            elif search_field == 'email':
                queryset = queryset.filter(email__icontains=search_value)
            elif search_field == 'phone':
                queryset = queryset.filter(phone__icontains=search_value)
            elif search_field == 'active':
                val = search_value.lower()
                if val in ['activo', 'activa', 'si', 'sí', '1', 'true']:
                    queryset = queryset.filter(is_active=True)
                elif val in ['inactivo', 'inactiva', 'no', '0', 'false']:
                    queryset = queryset.filter(is_active=False)
            else: # 'all'
                q_filters = Q(name__icontains=search_value) | \
                            Q(contact_name__icontains=search_value) | \
                            Q(email__icontains=search_value) | \
                            Q(phone__icontains=search_value)
                val = search_value.lower()
                if val in ['activo', 'activa', 'si', 'sí', '1', 'true']:
                    q_filters |= Q(is_active=True)
                elif val in ['inactivo', 'inactiva', 'no', '0', 'false']:
                    q_filters |= Q(is_active=False)
                queryset = queryset.filter(q_filters)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_field'] = self.request.GET.get('search_field', 'all').strip()
        context['search_value'] = self.request.GET.get('search_value', '').strip()
        return context

class SupplierCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'billing.add_supplier'
    model = Supplier; form_class = SupplierForm; template_name = 'billing/supplier_form.html'; success_url = reverse_lazy('billing:supplier_list')
class SupplierUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'billing.change_supplier'
    model = Supplier; form_class = SupplierForm; template_name = 'billing/supplier_form.html'; success_url = reverse_lazy('billing:supplier_list')
class SupplierDeleteView(LoginRequiredMixin, PermissionRequiredMixin, StaffRequiredMixin, DeleteView):
    permission_required = 'billing.delete_supplier'
    model = Supplier
    template_name = 'billing/supplier_confirm_delete.html'
    success_url = reverse_lazy('billing:supplier_list')
    staff_redirect_url = '/suppliers/'

# === PRODUCT (CBV) ===
class ProductListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'billing.view_product'
    model = Product
    template_name = 'billing/product_list.html'
    context_object_name = 'items'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset().select_related('brand', 'group').prefetch_related('suppliers')
        search_field = self.request.GET.get('search_field', 'all').strip()
        search_value = self.request.GET.get('search_value', '').strip()

        if search_value:
            from django.db.models import Q
            if search_field == 'name':
                queryset = queryset.filter(name__icontains=search_value)
            elif search_field == 'brand':
                queryset = queryset.filter(brand__name__icontains=search_value)
            elif search_field == 'group':
                queryset = queryset.filter(group__name__icontains=search_value)
            elif search_field == 'supplier':
                queryset = queryset.filter(suppliers__name__icontains=search_value).distinct()
            elif search_field == 'price':
                try:
                    price_val = Decimal(search_value)
                    queryset = queryset.filter(unit_price=price_val)
                except Exception:
                    queryset = queryset.none()
            elif search_field == 'stock':
                if search_value.isdigit():
                    queryset = queryset.filter(stock=int(search_value))
                else:
                    queryset = queryset.none()
            elif search_field == 'active':
                val = search_value.lower()
                if val in ['activo', 'activa', 'si', 'sí', '1', 'true']:
                    queryset = queryset.filter(is_active=True)
                elif val in ['inactivo', 'inactiva', 'no', '0', 'false']:
                    queryset = queryset.filter(is_active=False)
            else: # 'all'
                q_filters = Q(name__icontains=search_value) | \
                            Q(brand__name__icontains=search_value) | \
                            Q(group__name__icontains=search_value) | \
                            Q(suppliers__name__icontains=search_value)
                
                if search_value.isdigit():
                    q_filters |= Q(stock=int(search_value))
                try:
                    price_val = Decimal(search_value)
                    q_filters |= Q(unit_price=price_val)
                except Exception:
                    pass
                
                val = search_value.lower()
                if val in ['activo', 'activa', 'si', 'sí', '1', 'true']:
                    q_filters |= Q(is_active=True)
                elif val in ['inactivo', 'inactiva', 'no', '0', 'false']:
                    q_filters |= Q(is_active=False)
                
                queryset = queryset.filter(q_filters).distinct()
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_field'] = self.request.GET.get('search_field', 'all').strip()
        context['search_value'] = self.request.GET.get('search_value', '').strip()
        return context

class ProductCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'billing.add_product'
    model = Product; form_class = ProductForm; template_name = 'billing/product_form.html'; success_url = reverse_lazy('billing:product_list')
class ProductUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'billing.change_product'
    model = Product; form_class = ProductForm; template_name = 'billing/product_form.html'; success_url = reverse_lazy('billing:product_list')
class ProductDeleteView(LoginRequiredMixin, PermissionRequiredMixin, StaffRequiredMixin, DeleteView):
    permission_required = 'billing.delete_product'
    model = Product
    template_name = 'billing/product_confirm_delete.html'
    success_url = reverse_lazy('billing:product_list')
    staff_redirect_url = '/products/'

# === CUSTOMER (CBV) ===
class CustomerListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'billing.view_customer'
    model = Customer
    template_name = 'billing/customer_list.html'
    context_object_name = 'items'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()
        search_field = self.request.GET.get('search_field', 'all').strip()
        search_value = self.request.GET.get('search_value', '').strip()

        if search_value:
            from django.db.models import Q
            if search_field == 'name':
                queryset = queryset.filter(Q(first_name__icontains=search_value) | Q(last_name__icontains=search_value))
            elif search_field == 'dni':
                queryset = queryset.filter(dni__icontains=search_value)
            elif search_field == 'email':
                queryset = queryset.filter(email__icontains=search_value)
            elif search_field == 'phone':
                queryset = queryset.filter(phone__icontains=search_value)
            elif search_field == 'active':
                val = search_value.lower()
                if val in ['activo', 'activa', 'si', 'sí', '1', 'true']:
                    queryset = queryset.filter(is_active=True)
                elif val in ['inactivo', 'inactiva', 'no', '0', 'false']:
                    queryset = queryset.filter(is_active=False)
            else: # 'all'
                q_filters = Q(first_name__icontains=search_value) | \
                            Q(last_name__icontains=search_value) | \
                            Q(dni__icontains=search_value) | \
                            Q(email__icontains=search_value) | \
                            Q(phone__icontains=search_value)
                val = search_value.lower()
                if val in ['activo', 'activa', 'si', 'sí', '1', 'true']:
                    q_filters |= Q(is_active=True)
                elif val in ['inactivo', 'inactiva', 'no', '0', 'false']:
                    q_filters |= Q(is_active=False)
                queryset = queryset.filter(q_filters)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_field'] = self.request.GET.get('search_field', 'all').strip()
        context['search_value'] = self.request.GET.get('search_value', '').strip()
        return context

class CustomerCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'billing.add_customer'
    model = Customer; form_class = CustomerForm; template_name = 'billing/customer_form.html'; success_url = reverse_lazy('billing:customer_list')
class CustomerUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'billing.change_customer'
    model = Customer; form_class = CustomerForm; template_name = 'billing/customer_form.html'; success_url = reverse_lazy('billing:customer_list')
class CustomerDeleteView(LoginRequiredMixin, PermissionRequiredMixin, StaffRequiredMixin, DeleteView):
    permission_required = 'billing.delete_customer'
    model = Customer
    template_name = 'billing/customer_confirm_delete.html'
    success_url = reverse_lazy('billing:customer_list')
    staff_redirect_url = '/customers/'

# === INVOICE (CBV) ===
class InvoiceListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'billing.view_invoice'
    model = Invoice
    template_name = 'billing/invoice_list.html'
    context_object_name = 'items'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()
        search_field = self.request.GET.get('search_field', 'all').strip()
        search_value = self.request.GET.get('search_value', '').strip()

        if search_value:
            from django.db.models import Q
            if search_field == 'id':
                if search_value.isdigit():
                    queryset = queryset.filter(id=int(search_value))
                else:
                    queryset = queryset.none()
            elif search_field == 'customer_name':
                queryset = queryset.filter(
                    Q(customer__first_name__icontains=search_value) |
                    Q(customer__last_name__icontains=search_value)
                )
            elif search_field == 'customer_dni':
                queryset = queryset.filter(customer__dni__icontains=search_value)
            elif search_field == 'date':
                queryset = queryset.filter(invoice_date__date__icontains=search_value)
            elif search_field == 'total':
                try:
                    total_val = Decimal(search_value)
                    queryset = queryset.filter(total=total_val)
                except Exception:
                    queryset = queryset.none()
            else:  # 'all'
                q_filters = Q(customer__first_name__icontains=search_value) | \
                            Q(customer__last_name__icontains=search_value) | \
                            Q(customer__dni__icontains=search_value) | \
                            Q(invoice_date__date__icontains=search_value)
                
                if search_value.isdigit():
                    q_filters |= Q(id=int(search_value))
                
                try:
                    total_val = Decimal(search_value)
                    q_filters |= Q(total=total_val)
                except Exception:
                    pass
                
                queryset = queryset.filter(q_filters)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_field'] = self.request.GET.get('search_field', 'all').strip()
        context['search_value'] = self.request.GET.get('search_value', '').strip()
        return context

class InvoiceCreateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'billing.add_invoice'
    def get(self, request, *args, **kwargs):
        empresa = request.user.empresas.filter(is_active=True).first()
        if not empresa:
            empresa, _ = Empresa.objects.get_or_create(
                ruc="1756927560001",
                defaults={
                    'razon_social': 'Empresa Demo S.A.',
                    'nombre_comercial': 'Empresa Demo',
                    'dir_matriz': 'Quito, Ecuador',
                    'dir_establecimiento': 'Quito, Ecuador',
                    'obligado_contabilidad': False,
                    'codigo_establecimiento': '001',
                    'codigo_punto_emision': '001',
                    'secuencial_factura': 1,
                    'ambiente': '1',
                    'is_active': True
                }
            )
            empresa.usuarios.add(request.user)

        # Asegurar que Consumidor Final existe
        cf_obj, created = Customer.objects.get_or_create(
            dni="9999999999",
            defaults={
                'first_name': 'Consumidor',
                'last_name': 'Final',
                'is_active': True
            }
        )
        if created or not hasattr(cf_obj, 'profile') or not cf_obj.profile:
            CustomerProfile.objects.get_or_create(
                customer=cf_obj,
                defaults={
                    'taxpayer_type': 'final',
                    'payment_terms': 'cash',
                    'credit_limit': Decimal('0.00')
                }
            )

        customers = Customer.objects.filter(is_active=True).select_related('profile')
        products = Product.objects.filter(is_active=True).select_related('brand', 'group')
        
        customers_json = []
        for c in customers:
            taxpayer_type = 'final'
            payment_terms = 'cash'
            if hasattr(c, 'profile'):
                taxpayer_type = c.profile.taxpayer_type
                payment_terms = c.profile.payment_terms
            
            customers_json.append({
                'id': c.id,
                'dni': c.dni,
                'first_name': c.first_name,
                'last_name': c.last_name,
                'email': c.email or '',
                'phone': c.phone or '',
                'address': c.address or '',
                'taxpayer_type': taxpayer_type,
                'payment_terms': 'credit' if payment_terms.startswith('credit') else 'cash'
            })
            
        products_json = []
        for p in products:
            products_json.append({
                'id': p.id,
                'name': f"{p.name} ({p.brand.name})",
                'price': float(p.unit_price),
                'stock': p.stock
            })
            
        context = {
            'title': 'Crear Factura',
            'customers': customers,
            'products': products,
            'customers_json_str': json.dumps(customers_json),
            'products_json_str': json.dumps(products_json),
        }
        return render(request, 'billing/invoice_form.html', context)
        
    def post(self, request, *args, **kwargs):
        empresa = request.user.empresas.filter(is_active=True).first() or Empresa.objects.filter(is_active=True).first()
        customer_id = request.POST.get('customer_select', '').strip()
        customer_dni = request.POST.get('customer_dni', '').strip()
        customer_first_name = request.POST.get('customer_first_name', '').strip()
        customer_last_name = request.POST.get('customer_last_name', '').strip()
        customer_email = request.POST.get('customer_email', '').strip() or None
        customer_phone = request.POST.get('customer_phone', '').strip() or None
        customer_address = request.POST.get('customer_address', '').strip() or None
        customer_taxpayer_type = request.POST.get('customer_taxpayer_type', 'final')
        customer_payment_terms = request.POST.get('customer_payment_terms', 'cash')
        
        # Asegurar que Consumidor Final existe
        cf_obj, created = Customer.objects.get_or_create(
            dni="9999999999",
            defaults={
                'first_name': 'Consumidor',
                'last_name': 'Final',
                'is_active': True
            }
        )
        if created or not hasattr(cf_obj, 'profile') or not cf_obj.profile:
            CustomerProfile.objects.get_or_create(
                customer=cf_obj,
                defaults={
                    'taxpayer_type': 'final',
                    'payment_terms': 'cash',
                    'credit_limit': Decimal('0.00')
                }
            )

        # Si el cliente es Consumidor Final, forzar pago al contado
        if customer_dni == '9999999999' or (customer_id and customer_id != 'new' and Customer.objects.filter(id=customer_id, dni='9999999999').exists()):
            customer_payment_terms = 'cash'
        
        product_ids = request.POST.getlist('product[]')
        quantities = request.POST.getlist('quantity[]')
        prices = request.POST.getlist('price[]')
        
        errors = []
        if not customer_id:
            errors.append("Debe seleccionar un cliente o elegir '-- Nuevo Cliente --'.")
        
        if customer_id == 'new' or not customer_id:
            if not customer_dni:
                errors.append("El DNI/RUC es obligatorio para un nuevo cliente.")
            else:
                if not re.match(r'^\d+$', customer_dni):
                    errors.append("La identificación (DNI/RUC) solo debe contener números.")
                elif len(customer_dni) not in [10, 13]:
                    errors.append("La identificación debe tener exactamente 10 (DNI) o 13 (RUC) dígitos.")
            if not customer_first_name or not customer_last_name:
                errors.append("El Nombre y Apellido son obligatorios.")
            if customer_phone:
                phone_cleaned = re.sub(r'\s+|-', '', customer_phone)
                if not re.match(r'^\d+$', phone_cleaned):
                    errors.append("El teléfono solo debe contener números.")
                
        if not product_ids:
            errors.append("Debe agregar al menos un producto a la factura.")
            
        if len(product_ids) != len(quantities) or len(product_ids) != len(prices):
            errors.append("Datos de productos incorrectos.")
            
        items_to_save = []
        subtotal = Decimal('0.00')
        
        for idx in range(len(product_ids)):
            p_id = product_ids[idx]
            qty_str = quantities[idx]
            price_str = prices[idx]
            
            if not p_id or not qty_str or not price_str:
                errors.append(f"Fila {idx+1}: Complete todos los campos del producto.")
                continue
                
            try:
                qty = int(qty_str)
                price = Decimal(price_str)
            except (ValueError, TypeError):
                errors.append(f"Fila {idx+1}: Cantidad o precio no válidos.")
                continue
                
            if qty <= 0:
                errors.append(f"Fila {idx+1}: La cantidad debe ser mayor que cero.")
                continue
                
            if price < 0:
                errors.append(f"Fila {idx+1}: El precio no puede ser negativo.")
                continue
                
            product = Product.objects.filter(id=p_id).first()
            if not product:
                errors.append(f"Fila {idx+1}: El producto seleccionado no existe.")
                continue
                
            if product.stock < qty:
                errors.append(f"Stock insuficiente para {product.name}. Stock disponible: {product.stock}.")
                continue
                
            line_subtotal = qty * price
            subtotal += line_subtotal
            items_to_save.append({
                'product': product,
                'quantity': qty,
                'unit_price': price,
                'subtotal': line_subtotal
            })
            
        if errors:
            if request.GET.get('stream') == '1':
                def error_generator():
                    yield json.dumps({"step": "validation", "status": "error", "errors": errors}) + "\n"
                return StreamingHttpResponse(error_generator(), content_type='application/x-ndjson')
            else:
                for err in errors:
                    messages.error(request, err)
                return self._render_form_with_errors(request)
            
        if request.GET.get('stream') == '1':
            def stream_generator():
                nonlocal subtotal, customer_id, customer_dni, customer_first_name, customer_last_name
                nonlocal customer_email, customer_phone, customer_address, customer_taxpayer_type
                nonlocal customer_payment_terms, items_to_save, request
                
                local_customer = None
                local_invoice = None
                
                # Paso 1: Guardar Factura en Base de Datos
                yield json.dumps({"step": "db_save", "status": "running", "message": "Guardando factura en base de datos..."}) + "\n"
                try:
                    with transaction.atomic():
                        if customer_id and customer_id != 'new':
                            local_customer = Customer.objects.get(id=customer_id)
                            local_customer.first_name = customer_first_name
                            local_customer.last_name = customer_last_name
                            local_customer.email = customer_email
                            local_customer.phone = customer_phone
                            local_customer.address = customer_address
                            local_customer.save()
                        else:
                            local_customer = Customer.objects.filter(dni=customer_dni).first()
                            if local_customer:
                                local_customer.first_name = customer_first_name
                                local_customer.last_name = customer_last_name
                                local_customer.email = customer_email
                                local_customer.phone = customer_phone
                                local_customer.address = customer_address
                                local_customer.save()
                            else:
                                local_customer = Customer.objects.create(
                                    dni=customer_dni,
                                    first_name=customer_first_name,
                                    last_name=customer_last_name,
                                    email=customer_email,
                                    phone=customer_phone,
                                    address=customer_address
                                )
                        
                        profile, created = CustomerProfile.objects.get_or_create(customer=local_customer)
                        profile.taxpayer_type = customer_taxpayer_type
                        profile.payment_terms = customer_payment_terms
                        profile.save()
                        
                        tax = subtotal * Decimal('0.15')
                        total = subtotal + tax
                        
                        if customer_payment_terms == 'paypal':
                            metodo_pago = 'PAYPAL'
                            tipo_pago = 'CONTADO'
                            saldo = total
                            estado = 'PENDIENTE'
                        elif customer_payment_terms in ['credit', 'credit_15', 'credit_30', 'credit_60'] or customer_payment_terms.startswith('credit'):
                            metodo_pago = 'CREDITO'
                            tipo_pago = 'CREDITO'
                            saldo = total
                            estado = 'PENDIENTE'
                        else:
                            metodo_pago = 'EFECTIVO'
                            tipo_pago = 'CONTADO'
                            saldo = Decimal('0.00')
                            estado = 'PAGADA'

                        secuencial_factura_generada = str(empresa.secuencial_factura).zfill(9)
                        empresa.secuencial_factura += 1
                        empresa.save(update_fields=['secuencial_factura'])

                        local_invoice = Invoice.objects.create(
                            empresa=empresa,
                            customer=local_customer,
                            subtotal=subtotal,
                            tax=tax,
                            total=total,
                            is_active=True,
                            tipo_pago=tipo_pago,
                            metodo_pago=metodo_pago,
                            saldo=saldo,
                            estado=estado,
                            numero=f"{empresa.codigo_establecimiento}-{empresa.codigo_punto_emision}-{secuencial_factura_generada}"
                        )
                        
                        for item in items_to_save:
                            InvoiceDetail.objects.create(
                                invoice=local_invoice,
                                product=item['product'],
                                quantity=item['quantity'],
                                unit_price=item['unit_price'],
                                subtotal=item['subtotal']
                            )
                            item['product'].stock -= item['quantity']
                            item['product'].save()
                    
                    messages.success(request, f"Factura #{local_invoice.id} guardada con éxito.")
                    yield json.dumps({
                        "step": "db_save", 
                        "status": "success", 
                        "message": f"Factura #{local_invoice.id} guardada con éxito."
                    }) + "\n"
                except Exception as e:
                    yield json.dumps({"step": "db_save", "status": "error", "message": f"Error al guardar la factura: {str(e)}"}) + "\n"
                    return

                # Paso 2: Enviar por Correo Electrónico
                if local_customer.email:
                    yield json.dumps({
                        "step": "email_send", 
                        "status": "running", 
                        "message": f"Enviando correo a {local_customer.email}..."
                    }) + "\n"
                    try:
                        from django.core.mail import EmailMessage
                        pdf_data = generate_invoice_pdf_data(local_invoice)
                        email = EmailMessage(
                            subject=f'Factura #{local_invoice.id} - Sistema de Ventas',
                            body=(
                                f'Hola {local_customer.full_name},\n\n'
                                f'Se ha generado la factura #{local_invoice.id} de su compra realizada el {local_invoice.invoice_date.strftime("%d/%m/%Y")}.\n'
                                f'Adjunto a este correo encontrará el documento PDF con el detalle correspondiente.\n\n'
                                f'Detalles de facturación:\n'
                                f'- Subtotal: ${local_invoice.subtotal}\n'
                                f'- IVA (15%): ${local_invoice.tax}\n'
                                f'- Total Facturado: ${local_invoice.total}\n\n'
                                f'Agradecemos su preferencia.\n'
                                f'Atentamente,\nEl equipo de Ventas'
                            ),
                            from_email=None,
                            to=[local_customer.email]
                        )
                        email.attach(f'Factura_{local_invoice.id}.pdf', pdf_data, 'application/pdf')
                        email.send(fail_silently=False)
                        messages.success(request, f"Factura enviada automáticamente al correo: {local_customer.email}")
                        yield json.dumps({
                            "step": "email_send", 
                            "status": "success", 
                            "message": f"Factura enviada automáticamente al correo: {local_customer.email}"
                        }) + "\n"
                    except Exception as mail_err:
                        err_str = str(mail_err)
                        if "530" in err_str or "Authentication" in err_str:
                            err_str = "No se han configurado credenciales de correo en Vercel (Configura EMAIL_HOST_USER y EMAIL_HOST_PASSWORD en Vercel Settings)."
                        messages.warning(request, f"La factura fue guardada, pero no pudo ser enviada por correo: {err_str}")
                        yield json.dumps({
                            "step": "email_send", 
                            "status": "warning", 
                            "message": f"Fallo al enviar correo: {err_str}"
                        }) + "\n"
                else:
                    yield json.dumps({
                        "step": "email_send", 
                        "status": "skipped", 
                        "message": "Sin correo registrado para enviar."
                    }) + "\n"

                # Paso 3: Autorización del SRI (Ecuador)
                yield json.dumps({
                    "step": "sri_auth", 
                    "status": "running", 
                    "message": "Enviando comprobante electrónico al SRI..."
                }) + "\n"
                try:
                    import requests
                    import base64
                    import os
                    from django.conf import settings
                    
                    detalles_sri = []
                    for d in local_invoice.details.all():
                        detalles_sri.append({
                            "codigo": str(d.product.id),
                            "descripcion": d.product.name,
                            "cantidad": float(d.quantity),
                            "precio_unitario": float(d.unit_price),
                            "descuento": 0.0,
                            "precio_total": float(d.subtotal)
                        })
                    
                    dni = local_customer.dni
                    if dni in ("9999999999", "9999999999999"):
                        tipo_id = "07"
                    elif len(dni) == 13:
                        tipo_id = "04"
                    else:
                        tipo_id = "05"
                    
                    from django.utils import timezone
                    fecha_local = timezone.localtime(local_invoice.invoice_date).strftime("%d%m%Y")
                    
                    payload = {
                        "datos": {
                            "ambiente": empresa.ambiente,
                            "ruc_emisor": empresa.ruc,
                            "razon_social": empresa.razon_social,
                            "nombre_comercial": empresa.nombre_comercial,
                            "dir_matriz": empresa.dir_matriz,
                            "obligado_contabilidad": "SI" if empresa.obligado_contabilidad else "NO",
                            "serie": f"{empresa.codigo_establecimiento}{empresa.codigo_punto_emision}",
                            "secuencial": secuencial_factura_generada,
                            "fecha_emision": fecha_local,
                            "subtotal": float(local_invoice.subtotal),
                            "iva": float(local_invoice.tax),
                            "total": float(local_invoice.total),
                            "metodo_pago": local_invoice.metodo_pago,
                            "detalles": detalles_sri,
                            "cliente": {
                                "razon_social": local_customer.full_name,
                                "identificacion": dni,
                                "tipo_identificacion": tipo_id,
                                "email": local_customer.email,
                                "direccion": local_customer.address or empresa.dir_matriz
                            }
                        }
                    }
                    
                    response = requests.post(settings.SRI_MICROSERVICE_URL, json=payload, timeout=20)
                    if response.status_code == 200:
                        res_data = response.json()
                        clave = res_data.get("clave_acceso")
                        sec_sri = res_data.get("secuencial")
                        success = res_data.get("success")
                        
                        if clave:
                            local_invoice.clave_acceso = clave
                        if sec_sri:
                            local_invoice.numero = f"{empresa.codigo_establecimiento}-{empresa.codigo_punto_emision}-{sec_sri}"
                        
                        if success:
                            local_invoice.estado_sri = 'AUTORIZADO'
                            messages.success(request, f"Comprobante Autorizado por el SRI. Clave de Acceso: {clave}")
                            yield json.dumps({
                                "step": "sri_auth", 
                                "status": "success", 
                                "message": f"Comprobante Autorizado por el SRI. Clave de Acceso: {clave}"
                            }) + "\n"
                        else:
                            local_invoice.estado_sri = 'RECHAZADO'
                            err_msg = res_data.get("error", "Error devuelto por el SRI.")
                            messages.error(request, f"El SRI no autorizó la factura: {err_msg}")
                            
                            detalles = res_data.get("detalles") or []
                            errors_list = []
                            for d in detalles:
                                msg_info = f"- [{d.get('identificador', '')}] {d.get('mensaje', '')}"
                                if d.get('informacion_adicional'):
                                    msg_info += f" ({d.get('informacion_adicional')})"
                                messages.error(request, msg_info)
                                errors_list.append(msg_info)
                                
                            yield json.dumps({
                                "step": "sri_auth", 
                                "status": "warning", 
                                "message": f"El SRI no autorizó la factura: {err_msg}",
                                "errors": errors_list
                            }) + "\n"
                        
                        local_invoice.save(update_fields=['clave_acceso', 'estado_sri', 'numero'])
                    else:
                        try:
                            err_detail = response.json().get("detail", "Error interno en el microservicio.")
                        except Exception:
                            err_detail = "Error interno en el microservicio."
                        messages.error(request, f"Error del microservicio del SRI (HTTP {response.status_code}): {err_detail}")
                        yield json.dumps({
                            "step": "sri_auth", 
                            "status": "warning", 
                            "message": f"Error del microservicio del SRI: {err_detail}"
                        }) + "\n"
                except Exception as sri_err:
                    messages.warning(request, f"La factura se guardó, pero falló el envío al SRI: {str(sri_err)}")
                    yield json.dumps({
                        "step": "sri_auth", 
                        "status": "warning", 
                        "message": f"La factura se guardó, pero falló el envío al SRI: {str(sri_err)}"
                    }) + "\n"

                # Paso 4: Finalizado
                redirect_url = reverse('billing:invoice_list')
                if local_invoice.metodo_pago == 'PAYPAL':
                    redirect_url = reverse('billing:invoice_paypal_checkout', kwargs={'pk': local_invoice.pk})
                
                yield json.dumps({
                    "step": "finish", 
                    "status": "success", 
                    "message": "Factura generada y procesada exitosamente.", 
                    "redirect_url": redirect_url
                }) + "\n"

            response = StreamingHttpResponse(stream_generator(), content_type='application/x-ndjson')
            response['X-Accel-Buffering'] = 'no'
            return response

        # Comportamiento síncrono original (Fallback/Mantenimiento)
        try:
            with transaction.atomic():
                if customer_id and customer_id != 'new':
                    customer = Customer.objects.get(id=customer_id)
                    customer.first_name = customer_first_name
                    customer.last_name = customer_last_name
                    customer.email = customer_email
                    customer.phone = customer_phone
                    customer.address = customer_address
                    customer.save()
                else:
                    customer = Customer.objects.filter(dni=customer_dni).first()
                    if customer:
                        customer.first_name = customer_first_name
                        customer.last_name = customer_last_name
                        customer.email = customer_email
                        customer.phone = customer_phone
                        customer.address = customer_address
                        customer.save()
                    else:
                        customer = Customer.objects.create(
                            dni=customer_dni,
                            first_name=customer_first_name,
                            last_name=customer_last_name,
                            email=customer_email,
                            phone=customer_phone,
                            address=customer_address
                        )
                
                profile, created = CustomerProfile.objects.get_or_create(customer=customer)
                profile.taxpayer_type = customer_taxpayer_type
                profile.payment_terms = customer_payment_terms
                profile.save()
                
                tax = subtotal * Decimal('0.15')
                total = subtotal + tax
                
                if customer_payment_terms == 'paypal':
                    metodo_pago = 'PAYPAL'
                    tipo_pago = 'CONTADO'
                    saldo = total
                    estado = 'PENDIENTE'
                elif customer_payment_terms in ['credit', 'credit_15', 'credit_30', 'credit_60'] or customer_payment_terms.startswith('credit'):
                    metodo_pago = 'CREDITO'
                    tipo_pago = 'CREDITO'
                    saldo = total
                    estado = 'PENDIENTE'
                else:
                    metodo_pago = 'EFECTIVO'
                    tipo_pago = 'CONTADO'
                    saldo = Decimal('0.00')
                    estado = 'PAGADA'

                secuencial_factura_generada = str(empresa.secuencial_factura).zfill(9)
                empresa.secuencial_factura += 1
                empresa.save(update_fields=['secuencial_factura'])

                invoice = Invoice.objects.create(
                    empresa=empresa,
                    customer=customer,
                    subtotal=subtotal,
                    tax=tax,
                    total=total,
                    is_active=True,
                    tipo_pago=tipo_pago,
                    metodo_pago=metodo_pago,
                    saldo=saldo,
                    estado=estado,
                    numero=f"{empresa.codigo_establecimiento}-{empresa.codigo_punto_emision}-{secuencial_factura_generada}"
                )
                
                for item in items_to_save:
                    InvoiceDetail.objects.create(
                        invoice=invoice,
                        product=item['product'],
                        quantity=item['quantity'],
                        unit_price=item['unit_price'],
                        subtotal=item['subtotal']
                    )
                    item['product'].stock -= item['quantity']
                    item['product'].save()
                    
            # Enviar correo al cliente si tiene email registrado
            if customer.email:
                try:
                    from django.core.mail import EmailMessage
                    pdf_data = generate_invoice_pdf_data(invoice)
                    email = EmailMessage(
                        subject=f'Factura #{invoice.id} - Sistema de Ventas',
                        body=(
                            f'Hola {customer.full_name},\n\n'
                            f'Se ha generado la factura #{invoice.id} de su compra realizada el {invoice.invoice_date.strftime("%d/%m/%Y")}.\n'
                            f'Adjunto a este correo encontrará el documento PDF con el detalle correspondiente.\n\n'
                            f'Detalles de facturación:\n'
                            f'- Subtotal: ${invoice.subtotal}\n'
                            f'- IVA (15%): ${invoice.tax}\n'
                            f'- Total Facturado: ${invoice.total}\n\n'
                            f'Agradecemos su preferencia.\n'
                            f'Atentamente,\nEl equipo de Ventas'
                        ),
                        from_email=None,
                        to=[customer.email]
                    )
                    email.attach(f'Factura_{invoice.id}.pdf', pdf_data, 'application/pdf')
                    email.send(fail_silently=False)
                    messages.success(request, f"Factura enviada automáticamente al correo: {customer.email}")
                except Exception as mail_err:
                    messages.warning(request, f"La factura fue guardada, pero no pudo ser enviada por correo: {str(mail_err)}")

            # Enviar factura al Microservicio del SRI (Facturación Electrónica Ecuador)
            try:
                import requests
                import base64
                import os
                from django.conf import settings
                
                detalles_sri = []
                for d in invoice.details.all():
                    detalles_sri.append({
                        "codigo": str(d.product.id),
                        "descripcion": d.product.name,
                        "cantidad": float(d.quantity),
                        "precio_unitario": float(d.unit_price),
                        "descuento": 0.0,
                        "precio_total": float(d.subtotal)
                    })
                
                dni = customer.dni
                if dni in ("9999999999", "9999999999999"):
                    tipo_id = "07"
                elif len(dni) == 13:
                    tipo_id = "04"
                else:
                    tipo_id = "05"
                
                from django.utils import timezone
                fecha_local = timezone.localtime(invoice.invoice_date).strftime("%d%m%Y")
                
                payload = {
                    "datos": {
                        "ambiente": empresa.ambiente,
                        "ruc_emisor": empresa.ruc,
                        "razon_social": empresa.razon_social,
                        "nombre_comercial": empresa.nombre_comercial,
                        "dir_matriz": empresa.dir_matriz,
                        "obligado_contabilidad": "SI" if empresa.obligado_contabilidad else "NO",
                        "serie": f"{empresa.codigo_establecimiento}{empresa.codigo_punto_emision}",
                        "secuencial": secuencial_factura_generada,
                        "fecha_emision": fecha_local,
                        "subtotal": float(invoice.subtotal),
                        "iva": float(invoice.tax),
                        "total": float(invoice.total),
                        "metodo_pago": invoice.metodo_pago,
                        "detalles": detalles_sri,
                        "cliente": {
                            "razon_social": customer.full_name,
                            "identificacion": dni,
                            "tipo_identificacion": tipo_id,
                            "email": customer.email,
                            "direccion": customer.address or empresa.dir_matriz
                        }
                    }
                }
                
                # Consumir el microservicio
                response = requests.post(settings.SRI_MICROSERVICE_URL, json=payload, timeout=20)
                if response.status_code == 200:
                    res_data = response.json()
                    clave = res_data.get("clave_acceso")
                    sec_sri = res_data.get("secuencial")
                    success = res_data.get("success")
                    
                    if clave:
                        invoice.clave_acceso = clave
                    if sec_sri:
                        invoice.numero = f"{empresa.codigo_establecimiento}-{empresa.codigo_punto_emision}-{sec_sri}"
                    
                    if success:
                        invoice.estado_sri = 'AUTORIZADO'
                        messages.success(request, f"Comprobante Autorizado por el SRI. Clave de Acceso: {clave}")
                    else:
                        invoice.estado_sri = 'RECHAZADO'
                        err_msg = res_data.get("error", "Error devuelto por el SRI.")
                        messages.error(request, f"El SRI no autorizó la factura: {err_msg}")
                        
                        detalles = res_data.get("detalles") or []
                        for d in detalles:
                            msg_info = f"- [{d.get('identificador', '')}] {d.get('mensaje', '')}"
                            if d.get('informacion_adicional'):
                                msg_info += f" ({d.get('informacion_adicional')})"
                            messages.error(request, msg_info)
                    
                    invoice.save(update_fields=['clave_acceso', 'estado_sri', 'numero'])
                else:
                    try:
                        err_detail = response.json().get("detail", "Error interno en el microservicio.")
                    except Exception:
                        err_detail = "Error interno en el microservicio."
                    messages.error(request, f"Error del microservicio del SRI (HTTP {response.status_code}): {err_detail}")
            except Exception as sri_err:
                messages.warning(request, f"La factura se guardó, pero falló el envío al SRI: {str(sri_err)}")

            messages.success(request, f"Factura #{invoice.id} guardada con éxito.")
            if invoice.metodo_pago == 'PAYPAL':
                return redirect('billing:invoice_paypal_checkout', pk=invoice.pk)
            return redirect('billing:invoice_list')
            
        except Exception as e:
            messages.error(request, f"Error al guardar la factura: {str(e)}")
            return self._render_form_with_errors(request)
    def _render_form_with_errors(self, request):
        customers = Customer.objects.filter(is_active=True).select_related('profile')
        products = Product.objects.filter(is_active=True).select_related('brand', 'group')
        
        customers_json = []
        for c in customers:
            taxpayer_type = 'final'
            payment_terms = 'cash'
            if hasattr(c, 'profile'):
                taxpayer_type = c.profile.taxpayer_type
                payment_terms = c.profile.payment_terms
            
            customers_json.append({
                'id': c.id,
                'dni': c.dni,
                'first_name': c.first_name,
                'last_name': c.last_name,
                'email': c.email or '',
                'phone': c.phone or '',
                'address': c.address or '',
                'taxpayer_type': taxpayer_type,
                'payment_terms': 'credit' if payment_terms.startswith('credit') else 'cash'
            })
            
        products_json = []
        for p in products:
            products_json.append({
                'id': p.id,
                'name': f"{p.name} ({p.brand.name})",
                'price': float(p.unit_price),
                'stock': p.stock
            })
            
        product_ids = request.POST.getlist('product[]')
        quantities = request.POST.getlist('quantity[]')
        prices = request.POST.getlist('price[]')
        
        selected_items = []
        for idx in range(len(product_ids)):
            try:
                selected_items.append({
                    'product_id': int(product_ids[idx]) if product_ids[idx] else '',
                    'quantity': int(quantities[idx]) if quantities[idx] else 1,
                    'price': float(prices[idx]) if prices[idx] else 0.00
                })
            except (ValueError, TypeError):
                selected_items.append({
                    'product_id': product_ids[idx],
                    'quantity': quantities[idx],
                    'price': prices[idx]
                })

        posted_values = {
            'customer_select': request.POST.get('customer_select', ''),
            'customer_dni': request.POST.get('customer_dni', ''),
            'customer_first_name': request.POST.get('customer_first_name', ''),
            'customer_last_name': request.POST.get('customer_last_name', ''),
            'customer_email': request.POST.get('customer_email', '') or '',
            'customer_phone': request.POST.get('customer_phone', '') or '',
            'customer_address': request.POST.get('customer_address', '') or '',
            'customer_taxpayer_type': request.POST.get('customer_taxpayer_type', 'final'),
            'customer_payment_terms': request.POST.get('customer_payment_terms', 'cash'),
        }

        context = {
            'title': 'Crear Factura',
            'customers': customers,
            'products': products,
            'customers_json_str': json.dumps(customers_json),
            'products_json_str': json.dumps(products_json),
            'posted_values_json': json.dumps(posted_values),
            'selected_items_json': json.dumps(selected_items),
        }
        return render(request, 'billing/invoice_form.html', context)
class InvoiceDeleteView(LoginRequiredMixin, PermissionRequiredMixin, StaffRequiredMixin, DeleteView):
    permission_required = 'billing.delete_invoice'
    model = Invoice
    template_name = 'billing/invoice_confirm_delete.html'
    success_url = reverse_lazy('billing:invoice_list')
    staff_redirect_url = '/invoices/'

    def post(self, request, *args, **kwargs):
        from django.db.models import ProtectedError
        from django.contrib import messages
        from django.shortcuts import redirect
        try:
            return super().post(request, *args, **kwargs)
        except ProtectedError as e:
            related_objects = e.protected_objects
            abonos_info = []
            for obj in related_objects:
                if hasattr(obj, 'fecha') and hasattr(obj, 'valor'):
                    abonos_info.append(f"Abono #{obj.id} por ${obj.valor} ({obj.fecha})")
                else:
                    abonos_info.append(str(obj))
                    
            error_msg = "No se puede eliminar la factura porque está protegida y vinculada a transacciones existentes. "
            if abonos_info:
                error_msg += f"Cobros/Abonos asociados: {', '.join(abonos_info)}. "
            error_msg += "Por favor, elimine primero estos cobros o abonos en el módulo financiero antes de eliminar la factura."
            
            messages.error(request, error_msg)
            return redirect('billing:invoice_list')


# === REPORTES DE PRODUCTOS (Excel / PDF) ===
@login_required
@permission_required('billing.view_product')
def product_report_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    ws.merge_cells("A1:F1")
    ws["A1"] = "REPORTE GENERAL DE PRODUCTOS"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    headers = ["Nombre del Producto", "Marca", "Categoría/Grupo", "Precio Unitario (USD)", "Stock", "Balance (USD)"]
    ws.append([])
    ws.append(headers)
    
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    for col in range(1, 7):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    ws.row_dimensions[3].height = 25
    
    products = Product.objects.all().order_by('name')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for p in products:
        row = [
            p.name,
            p.brand.name if p.brand else "-",
            p.group.name if p.group else "-",
            float(p.unit_price),
            p.stock,
            float(p.balance)
        ]
        ws.append(row)
        curr_row = ws.max_row
        ws.row_dimensions[curr_row].height = 20
        
        ws.cell(row=curr_row, column=1).alignment = left_align
        ws.cell(row=curr_row, column=2).alignment = left_align
        ws.cell(row=curr_row, column=3).alignment = left_align
        
        price_cell = ws.cell(row=curr_row, column=4)
        price_cell.alignment = right_align
        price_cell.number_format = '$#,##0.00'
        
        stock_cell = ws.cell(row=curr_row, column=5)
        stock_cell.alignment = right_align
        stock_cell.number_format = '#,##0'

        balance_cell = ws.cell(row=curr_row, column=6)
        balance_cell.alignment = right_align
        balance_cell.number_format = '$#,##0.00'
        
        for col in range(1, 7):
            c = ws.cell(row=curr_row, column=col)
            c.font = Font(name="Calibri", size=11)
            c.border = thin_border
            
    ws.auto_filter.ref = f"A3:F{ws.max_row}"
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            if cell.column in [4, 6] and isinstance(cell.value, (int, float)):
                val_str = f"${cell.value:,.2f}"
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Reporte_Productos.xlsx"'
    wb.save(response)
    return response

@login_required
@permission_required('billing.view_product')
def product_report_pdf(request):
    import io
    import datetime
    from django.http import FileResponse
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    set_current_report_title("Sistema de Ventas - Reporte de Productos")
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=54,
        bottomMargin=72
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=22,
        textColor=colors.HexColor("#1F4E78"),
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor("#595959"),
        spaceAfter=25
    )
    cell_header_style = ParagraphStyle(
        'CellHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.white,
        alignment=1
    )
    cell_text_left = ParagraphStyle(
        'CellTextLeft',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor("#262626")
    )
    cell_text_right = ParagraphStyle(
        'CellTextRight',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor("#262626"),
        alignment=2
    )
    
    now_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    story.append(Paragraph("REPORTE GENERAL DE PRODUCTOS", title_style))
    story.append(Paragraph(f"Fecha de Generación: {now_str} | Total de productos registrados en base de datos.", subtitle_style))
    
    products = Product.objects.all().order_by('name')
    
    data = [
        [
            Paragraph("Nombre del Producto", cell_header_style),
            Paragraph("Marca", cell_header_style),
            Paragraph("Categoría", cell_header_style),
            Paragraph("Precio (USD)", cell_header_style),
            Paragraph("Stock", cell_header_style),
            Paragraph("Balance (USD)", cell_header_style)
        ]
    ]
    
    for p in products:
        data.append([
            Paragraph(p.name, cell_text_left),
            Paragraph(p.brand.name if p.brand else "-", cell_text_left),
            Paragraph(p.group.name if p.group else "-", cell_text_left),
            Paragraph(f"${p.unit_price:,.2f}", cell_text_right),
            Paragraph(f"{p.stock:,}", cell_text_right),
            Paragraph(f"${p.balance:,.2f}", cell_text_right)
        ])
        
    col_widths = [140, 80, 80, 80, 80, 80]
    
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2F5597")),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#F2F4F7"), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))
    
    story.append(t)
    doc.build(story, canvasmaker=NumberedCanvas)
    
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename="Reporte_Productos.pdf")


# === REPORTES DE MARCAS (Excel / PDF) ===
@login_required
@permission_required('billing.view_brand')
def brand_report_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marcas"
    ws.merge_cells("A1:D1")
    ws["A1"] = "REPORTE GENERAL DE MARCAS"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    headers = ["ID", "Nombre de Marca", "Descripción", "Estado"]
    ws.append([])
    ws.append(headers)

    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col in range(1, 5):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 25

    brands = Brand.objects.all().order_by('name')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for b in brands:
        row = [
            b.id,
            b.name,
            b.description or "",
            "Activo" if b.is_active else "Inactivo"
        ]
        ws.append(row)
        curr_row = ws.max_row
        ws.row_dimensions[curr_row].height = 20
        ws.cell(row=curr_row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=curr_row, column=4).alignment = Alignment(horizontal="center")
        for col in range(1, 5):
            c = ws.cell(row=curr_row, column=col)
            c.font = Font(name="Calibri", size=11)
            c.border = thin_border

    ws.auto_filter.ref = f"A3:D{ws.max_row}"
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Reporte_Marcas.xlsx"'
    wb.save(response)
    return response


@login_required
@permission_required('billing.view_brand')
def brand_report_pdf(request):
    import io
    import datetime
    from django.http import FileResponse
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    set_current_report_title("Sistema de Ventas - Reporte de Marcas")
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=54,
        bottomMargin=72
    )

    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=22,
        textColor=colors.HexColor("#1F4E78"),
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor("#595959"),
        spaceAfter=25
    )
    cell_header_style = ParagraphStyle(
        'CellHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.white,
        alignment=1
    )
    cell_text_left = ParagraphStyle(
        'CellTextLeft',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor("#262626")
    )
    cell_text_center = ParagraphStyle(
        'CellTextCenter',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor("#262626"),
        alignment=1
    )

    now_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    story.append(Paragraph("REPORTE GENERAL DE MARCAS", title_style))
    story.append(Paragraph(f"Fecha de Generación: {now_str} | Total de marcas registradas.", subtitle_style))

    brands = Brand.objects.all().order_by('name')

    data = [
        [
            Paragraph("ID", cell_header_style),
            Paragraph("Nombre de Marca", cell_header_style),
            Paragraph("Descripción", cell_header_style),
            Paragraph("Estado", cell_header_style)
        ]
    ]

    for b in brands:
        data.append([
            Paragraph(f"#{b.id}", cell_text_center),
            Paragraph(b.name, cell_text_left),
            Paragraph(b.description or "", cell_text_left),
            Paragraph("Activo" if b.is_active else "Inactivo", cell_text_center)
        ])

    col_widths = [50, 150, 240, 100]

    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2F5597")),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#F2F4F7"), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))

    story.append(t)
    doc.build(story, canvasmaker=NumberedCanvas)

    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename="Reporte_Marcas.pdf")


# === REPORTES DE CATEGORIAS / GRUPOS (Excel / PDF) ===
@login_required
@permission_required('billing.view_productgroup')
def productgroup_report_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grupos"
    ws.merge_cells("A1:C1")
    ws["A1"] = "REPORTE GENERAL DE GRUPOS DE PRODUCTOS"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    headers = ["ID", "Nombre de Grupo", "Estado"]
    ws.append([])
    ws.append(headers)

    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col in range(1, 4):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 25

    groups = ProductGroup.objects.all().order_by('name')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for g in groups:
        row = [
            g.id,
            g.name,
            "Activo" if g.is_active else "Inactivo"
        ]
        ws.append(row)
        curr_row = ws.max_row
        ws.row_dimensions[curr_row].height = 20
        ws.cell(row=curr_row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=curr_row, column=3).alignment = Alignment(horizontal="center")
        for col in range(1, 4):
            c = ws.cell(row=curr_row, column=col)
            c.font = Font(name="Calibri", size=11)
            c.border = thin_border

    ws.auto_filter.ref = f"A3:C{ws.max_row}"
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Reporte_Grupos.xlsx"'
    wb.save(response)
    return response


@login_required
@permission_required('billing.view_productgroup')
def productgroup_report_pdf(request):
    import io
    import datetime
    from django.http import FileResponse
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    set_current_report_title("Sistema de Ventas - Reporte de Categorías")
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=54,
        bottomMargin=72
    )

    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=22,
        textColor=colors.HexColor("#1F4E78"),
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor("#595959"),
        spaceAfter=25
    )
    cell_header_style = ParagraphStyle(
        'CellHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.white,
        alignment=1
    )
    cell_text_left = ParagraphStyle(
        'CellTextLeft',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor("#262626")
    )
    cell_text_center = ParagraphStyle(
        'CellTextCenter',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor("#262626"),
        alignment=1
    )

    now_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    story.append(Paragraph("REPORTE GENERAL DE GRUPOS DE PRODUCTOS", title_style))
    story.append(Paragraph(f"Fecha de Generación: {now_str} | Total de categorías/grupos registrados.", subtitle_style))

    groups = ProductGroup.objects.all().order_by('name')

    data = [
        [
            Paragraph("ID", cell_header_style),
            Paragraph("Nombre de Grupo", cell_header_style),
            Paragraph("Estado", cell_header_style)
        ]
    ]

    for g in groups:
        data.append([
            Paragraph(f"#{g.id}", cell_text_center),
            Paragraph(g.name, cell_text_left),
            Paragraph("Activo" if g.is_active else "Inactivo", cell_text_center)
        ])

    col_widths = [80, 320, 140]

    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2F5597")),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#F2F4F7"), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))

    story.append(t)
    doc.build(story, canvasmaker=NumberedCanvas)

    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename="Reporte_Grupos.pdf")


# === REPORTES DE PROVEEDORES (Excel / PDF) ===
@login_required
@permission_required('billing.view_supplier')
def supplier_report_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    ws.merge_cells("A1:G1")
    ws["A1"] = "REPORTE GENERAL DE PROVEEDORES"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    headers = ["ID", "Razón Social", "Contacto", "Email", "Teléfono", "Dirección", "Estado"]
    ws.append([])
    ws.append(headers)

    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col in range(1, 8):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 25

    suppliers = Supplier.objects.all().order_by('name')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for s in suppliers:
        row = [
            s.id,
            s.name,
            s.contact_name or "",
            s.email or "",
            s.phone or "",
            s.address or "",
            "Activo" if s.is_active else "Inactivo"
        ]
        ws.append(row)
        curr_row = ws.max_row
        ws.row_dimensions[curr_row].height = 20
        ws.cell(row=curr_row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=curr_row, column=7).alignment = Alignment(horizontal="center")
        for col in range(1, 8):
            c = ws.cell(row=curr_row, column=col)
            c.font = Font(name="Calibri", size=11)
            c.border = thin_border

    ws.auto_filter.ref = f"A3:G{ws.max_row}"
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Reporte_Proveedores.xlsx"'
    wb.save(response)
    return response


@login_required
@permission_required('billing.view_supplier')
def supplier_report_pdf(request):
    import io
    import datetime
    from django.http import FileResponse
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    set_current_report_title("Sistema de Ventas - Reporte de Proveedores")
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=54,
        bottomMargin=72
    )

    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=22,
        textColor=colors.HexColor("#1F4E78"),
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor("#595959"),
        spaceAfter=25
    )
    cell_header_style = ParagraphStyle(
        'CellHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
        alignment=1
    )
    cell_text_left = ParagraphStyle(
        'CellTextLeft',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor("#262626")
    )
    cell_text_center = ParagraphStyle(
        'CellTextCenter',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor("#262626"),
        alignment=1
    )

    now_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    story.append(Paragraph("REPORTE GENERAL DE PROVEEDORES", title_style))
    story.append(Paragraph(f"Fecha de Generación: {now_str} | Total de proveedores registrados.", subtitle_style))

    suppliers = Supplier.objects.all().order_by('name')

    data = [
        [
            Paragraph("ID", cell_header_style),
            Paragraph("Razón Social", cell_header_style),
            Paragraph("Contacto", cell_header_style),
            Paragraph("Email", cell_header_style),
            Paragraph("Teléfono", cell_header_style),
            Paragraph("Estado", cell_header_style)
        ]
    ]

    for s in suppliers:
        data.append([
            Paragraph(f"#{s.id}", cell_text_center),
            Paragraph(s.name, cell_text_left),
            Paragraph(s.contact_name or "", cell_text_left),
            Paragraph(s.email or "", cell_text_left),
            Paragraph(s.phone or "", cell_text_center),
            Paragraph("Activo" if s.is_active else "Inactivo", cell_text_center)
        ])

    col_widths = [40, 140, 100, 130, 80, 50]

    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2F5597")),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#F2F4F7"), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))

    story.append(t)
    doc.build(story, canvasmaker=NumberedCanvas)

    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename="Reporte_Proveedores.pdf")


# === REPORTES DE CLIENTES (Excel / PDF) ===
@login_required
@permission_required('billing.view_customer')
def customer_report_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.merge_cells("A1:G1")
    ws["A1"] = "REPORTE GENERAL DE CLIENTES"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    headers = ["ID", "DNI/RUC", "Nombre Completo", "Email", "Teléfono", "Dirección", "Estado"]
    ws.append([])
    ws.append(headers)

    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col in range(1, 8):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 25

    customers = Customer.objects.all().order_by('last_name', 'first_name')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for c in customers:
        row = [
            c.id,
            c.dni,
            c.full_name,
            c.email or "",
            c.phone or "",
            c.address or "",
            "Activo" if c.is_active else "Inactivo"
        ]
        ws.append(row)
        curr_row = ws.max_row
        ws.row_dimensions[curr_row].height = 20
        ws.cell(row=curr_row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=curr_row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=curr_row, column=7).alignment = Alignment(horizontal="center")
        for col in range(1, 8):
            cell_obj = ws.cell(row=curr_row, column=col)
            cell_obj.font = Font(name="Calibri", size=11)
            cell_obj.border = thin_border

    ws.auto_filter.ref = f"A3:G{ws.max_row}"
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Reporte_Clientes.xlsx"'
    wb.save(response)
    return response


@login_required
@permission_required('billing.view_customer')
def customer_report_pdf(request):
    import io
    import datetime
    from django.http import FileResponse
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    set_current_report_title("Sistema de Ventas - Reporte de Clientes")
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=54,
        bottomMargin=72
    )

    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=22,
        textColor=colors.HexColor("#1F4E78"),
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor("#595959"),
        spaceAfter=25
    )
    cell_header_style = ParagraphStyle(
        'CellHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
        alignment=1
    )
    cell_text_left = ParagraphStyle(
        'CellTextLeft',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor("#262626")
    )
    cell_text_center = ParagraphStyle(
        'CellTextCenter',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor("#262626"),
        alignment=1
    )

    now_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    story.append(Paragraph("REPORTE GENERAL DE CLIENTES", title_style))
    story.append(Paragraph(f"Fecha de Generación: {now_str} | Total de clientes registrados.", subtitle_style))

    customers = Customer.objects.all().order_by('last_name', 'first_name')

    data = [
        [
            Paragraph("DNI/RUC", cell_header_style),
            Paragraph("Nombre Completo", cell_header_style),
            Paragraph("Email", cell_header_style),
            Paragraph("Teléfono", cell_header_style),
            Paragraph("Dirección", cell_header_style),
            Paragraph("Estado", cell_header_style)
        ]
    ]

    for c in customers:
        data.append([
            Paragraph(c.dni, cell_text_center),
            Paragraph(c.full_name, cell_text_left),
            Paragraph(c.email or "", cell_text_left),
            Paragraph(c.phone or "", cell_text_center),
            Paragraph(c.address or "", cell_text_left),
            Paragraph("Activo" if c.is_active else "Inactivo", cell_text_center)
        ])

    col_widths = [90, 140, 130, 80, 55, 45]

    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2F5597")),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#F2F4F7"), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))

    story.append(t)
    doc.build(story, canvasmaker=NumberedCanvas)

    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename="Reporte_Clientes.pdf")


# === REPORTES DE FACTURAS (Excel / PDF) ===
@login_required
@permission_required('billing.view_invoice')
def invoice_report_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas"
    ws.merge_cells("A1:G1")
    ws["A1"] = "REPORTE GENERAL DE FACTURAS EMITIDAS"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    headers = ["ID Factura", "Cliente", "Fecha de Emisión", "Subtotal (USD)", "IVA (15%)", "Total (USD)", "Estado"]
    ws.append([])
    ws.append(headers)

    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col in range(1, 8):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 25

    invoices = Invoice.objects.select_related('customer').all().order_by('-invoice_date')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for i in invoices:
        row = [
            i.id,
            i.customer.full_name,
            i.invoice_date.strftime("%d/%m/%Y %H:%M"),
            float(i.subtotal),
            float(i.tax),
            float(i.total),
            "Activo" if i.is_active else "Anulado"
        ]
        ws.append(row)
        curr_row = ws.max_row
        ws.row_dimensions[curr_row].height = 20
        ws.cell(row=curr_row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=curr_row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=curr_row, column=7).alignment = Alignment(horizontal="center")
        
        ws.cell(row=curr_row, column=4).number_format = '$#,##0.00'
        ws.cell(row=curr_row, column=5).number_format = '$#,##0.00'
        ws.cell(row=curr_row, column=6).number_format = '$#,##0.00'
        ws.cell(row=curr_row, column=4).alignment = Alignment(horizontal="right")
        ws.cell(row=curr_row, column=5).alignment = Alignment(horizontal="right")
        ws.cell(row=curr_row, column=6).alignment = Alignment(horizontal="right")

        for col in range(1, 8):
            c = ws.cell(row=curr_row, column=col)
            c.font = Font(name="Calibri", size=11)
            c.border = thin_border

    ws.auto_filter.ref = f"A3:G{ws.max_row}"
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            if cell.column in [4, 5, 6] and isinstance(cell.value, (int, float)):
                val_str = f"${cell.value:,.2f}"
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Reporte_Facturas.xlsx"'
    wb.save(response)
    return response


@login_required
@permission_required('billing.view_invoice')
def invoice_report_pdf(request):
    import io
    import datetime
    from django.http import FileResponse
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    set_current_report_title("Sistema de Ventas - Reporte de Facturas")
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=54,
        bottomMargin=72
    )

    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=22,
        textColor=colors.HexColor("#1F4E78"),
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor("#595959"),
        spaceAfter=25
    )
    cell_header_style = ParagraphStyle(
        'CellHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
        alignment=1
    )
    cell_text_left = ParagraphStyle(
        'CellTextLeft',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor("#262626")
    )
    cell_text_center = ParagraphStyle(
        'CellTextCenter',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor("#262626"),
        alignment=1
    )
    cell_text_right = ParagraphStyle(
        'CellTextRight',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor("#262626"),
        alignment=2
    )

    now_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    story.append(Paragraph("REPORTE GENERAL DE FACTURAS", title_style))
    story.append(Paragraph(f"Fecha de Generación: {now_str} | Historial de facturas emitidas.", subtitle_style))

    invoices = Invoice.objects.select_related('customer').all().order_by('-invoice_date')

    data = [
        [
            Paragraph("Factura #", cell_header_style),
            Paragraph("Cliente", cell_header_style),
            Paragraph("Fecha de Emisión", cell_header_style),
            Paragraph("Subtotal", cell_header_style),
            Paragraph("IVA (15%)", cell_header_style),
            Paragraph("Total", cell_header_style),
            Paragraph("Estado", cell_header_style)
        ]
    ]

    for i in invoices:
        data.append([
            Paragraph(f"#{i.id}", cell_text_center),
            Paragraph(i.customer.full_name, cell_text_left),
            Paragraph(i.invoice_date.strftime("%d/%m/%Y %H:%M"), cell_text_center),
            Paragraph(f"${i.subtotal:,.2f}", cell_text_right),
            Paragraph(f"${i.tax:,.2f}", cell_text_right),
            Paragraph(f"${i.total:,.2f}", cell_text_right),
            Paragraph("Activo" if i.is_active else "Anulado", cell_text_center)
        ])

    col_widths = [55, 140, 95, 90, 50, 50, 60]

    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2F5597")),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#F2F4F7"), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))

    story.append(t)
    doc.build(story, canvasmaker=NumberedCanvas)

    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename="Reporte_Facturas.pdf")


def generate_invoice_pdf_data(invoice):
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=54,
        bottomMargin=54
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Estilos Personalizados
    title_style = ParagraphStyle(
        'InvoiceTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor("#1F4E78"),
        spaceAfter=5
    )
    normal_style = ParagraphStyle(
        'InvoiceNormal',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor("#262626"),
        spaceAfter=3
    )
    bold_style = ParagraphStyle(
        'InvoiceBold',
        parent=normal_style,
        fontName='Helvetica-Bold'
    )
    
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.white,
        alignment=1
    )
    
    cell_left = ParagraphStyle(
        'CellLeft',
        parent=normal_style,
        fontSize=9
    )
    cell_center = ParagraphStyle(
        'CellCenter',
        parent=normal_style,
        fontSize=9,
        alignment=1
    )
    cell_right = ParagraphStyle(
        'CellRight',
        parent=normal_style,
        fontSize=9,
        alignment=2
    )

    # Encabezado
    story.append(Paragraph("FACTURA DE VENTA", title_style))
    story.append(Paragraph(f"Factura Nº: {invoice.id}", bold_style))
    story.append(Paragraph(f"Fecha: {invoice.invoice_date.strftime('%d/%m/%Y %H:%M')}", normal_style))
    story.append(Spacer(1, 15))
    
    # Tabla de Datos Emisor / Receptor
    info_data = [
        [
            Paragraph("<b>EMISOR</b><br/><b>Empresa:</b> Sistema de Ventas<br/><b>Email:</b> ventas@sistema.com", normal_style),
            Paragraph(f"<b>CLIENTE</b><br/><b>Nombre:</b> {invoice.customer.full_name}<br/><b>DNI/RUC:</b> {invoice.customer.dni}<br/><b>Email:</b> {invoice.customer.email or '-'}<br/><b>Tlf:</b> {invoice.customer.phone or '-'}", normal_style)
        ]
    ]
    info_table = Table(info_data, colWidths=[250, 250])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('PADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 20))
    
    # Tabla de Productos
    table_data = [
        [
            Paragraph("Producto", header_style),
            Paragraph("Cant.", header_style),
            Paragraph("Precio Unitario", header_style),
            Paragraph("Subtotal", header_style)
        ]
    ]
    
    for detail in invoice.details.all():
        table_data.append([
            Paragraph(detail.product.name, cell_left),
            Paragraph(str(detail.quantity), cell_center),
            Paragraph(f"${detail.unit_price:,.2f}", cell_right),
            Paragraph(f"${detail.subtotal:,.2f}", cell_right)
        ])
        
    items_table = Table(table_data, colWidths=[260, 60, 90, 90])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1F4E78")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#D9D9D9")),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(items_table)
    story.append(Spacer(1, 15))
    
    # Tabla de Totales
    totals_data = [
        [Paragraph("", cell_left), Paragraph("Subtotal:", cell_right), Paragraph(f"${invoice.subtotal:,.2f}", cell_right)],
        [Paragraph("", cell_left), Paragraph("IVA (15%):", cell_right), Paragraph(f"${invoice.tax:,.2f}", cell_right)],
        [Paragraph("", cell_left), Paragraph("TOTAL:", ParagraphStyle('TBold', parent=cell_right, fontName='Helvetica-Bold')), Paragraph(f"${invoice.total:,.2f}", ParagraphStyle('TBold2', parent=cell_right, fontName='Helvetica-Bold'))]
    ]
    totals_table = Table(totals_data, colWidths=[320, 90, 90])
    totals_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LINEABOVE', (1,2), (2,2), 1, colors.HexColor("#1F4E78")),
    ]))
    story.append(totals_table)
    
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


@login_required
@permission_required('billing.view_invoice')
def invoice_pdf(request, pk):
    from django.http import HttpResponse
    invoice = get_object_or_404(Invoice, pk=pk)
    pdf_data = generate_invoice_pdf_data(invoice)
    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Factura_{invoice.id}.pdf"'
    return response


@login_required
@permission_required('billing.view_invoice')
def invoice_paypal_checkout(request, pk):
    from django.conf import settings
    invoice = get_object_or_404(Invoice, pk=pk)
    if invoice.estado == 'PAGADA':
        messages.warning(request, f"La factura #{invoice.id} ya se encuentra pagada.")
        return redirect('billing:invoice_detail', pk=invoice.pk)
    
    context = {
        'invoice': invoice,
        'invoice_total_paypal': format(invoice.total, '.2f'),
        'paypal_client_id': settings.PAYPAL_CLIENT_ID,
        'title': f"Pago PayPal - Factura #{invoice.id}"
    }
    return render(request, 'billing/invoice_paypal_checkout.html', context)


@login_required
@permission_required('billing.change_invoice')
def invoice_paypal_capture(request, pk):
    from django.http import JsonResponse
    import json
    from pagos.models import CobroFactura
    from django.utils import timezone
    
    invoice = get_object_or_404(Invoice, pk=pk)
    if invoice.estado == 'PAGADA':
        return JsonResponse({'status': 'error', 'message': 'La factura ya se encuentra pagada.'}, status=400)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            order_id = data.get('orderID')
            payer_id = data.get('payerID')
            
            if not order_id:
                return JsonResponse({'status': 'error', 'message': 'ID de Orden de PayPal faltante.'}, status=400)
            
            with transaction.atomic():
                invoice = Invoice.objects.select_for_update().get(pk=pk)
                if invoice.estado == 'PAGADA':
                    return JsonResponse({'status': 'error', 'message': 'La factura ya se encuentra pagada.'}, status=400)
                
                monto_pagado = invoice.saldo
                
                # Registrar en CobroFactura (la cual decrementa el saldo y actualiza el estado de la factura automáticamente)
                CobroFactura.objects.create(
                    factura=invoice,
                    fecha=timezone.now(),
                    valor=monto_pagado,
                    observacion=f"Pago completo realizado a través de PayPal (Orden ID: {order_id}, Payer ID: {payer_id})"
                )
                
            messages.success(request, f"Pago con PayPal de ${monto_pagado} capturado exitosamente para la Factura #{invoice.id}.")
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
            
    return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)


@login_required
def profile_edit(request):
    from .forms import UserProfileForm
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Tu perfil ha sido actualizado con éxito!')
            return redirect('billing:home')
    else:
        form = UserProfileForm(instance=request.user)
    return render(request, 'billing/profile_edit.html', {'form': form})


