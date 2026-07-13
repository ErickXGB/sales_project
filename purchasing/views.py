from shared.decorators import permission_required
import json
import re
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import F, Avg, Q
from decimal import Decimal
from billing.models import Supplier, Product
from .models import Purchase, PurchaseDetail

@login_required
@permission_required('purchasing.view_purchase')
def purchase_list(request):
    """Lista todas las compras con búsqueda por campo (igual que product_list)."""
    purchases = Purchase.objects.select_related('supplier').all()

    search_field = request.GET.get('search_field', 'all').strip()
    search_value = request.GET.get('search_value', '').strip()

    if search_value:
        if search_field == 'supplier' or search_field == 'all':
            if search_field == 'all':
                purchases = purchases.filter(
                    Q(supplier__name__icontains=search_value) |
                    Q(document_number__icontains=search_value) |
                    Q(purchase_date__icontains=search_value) |
                    Q(total__icontains=search_value)
                )
            else:
                purchases = purchases.filter(supplier__name__icontains=search_value)
        elif search_field == 'document':
            purchases = purchases.filter(document_number__icontains=search_value)
        elif search_field == 'date':
            purchases = purchases.filter(purchase_date__date__icontains=search_value)
        elif search_field == 'total':
            purchases = purchases.filter(total__icontains=search_value)
        elif search_field == 'active':
            is_active = search_value.lower() in ['activo', 'si', 'sí', '1', 'true', 'yes']
            purchases = purchases.filter(is_active=is_active)

    context = {
        'items': purchases,
        'search_field': search_field,
        'search_value': search_value,
    }
    return render(request, 'purchasing/purchase_list.html', context)

@login_required
@permission_required('purchasing.add_purchase')
def purchase_create(request):
    """Crea una compra con sus detalles usando entrada dinámica (JS)."""
    if request.method == 'POST':
        supplier_select = request.POST.get('supplier_select', '').strip()
        supplier_name = request.POST.get('supplier_name', '').strip()
        supplier_contact_name = request.POST.get('supplier_contact_name', '').strip()
        supplier_email = request.POST.get('supplier_email', '').strip() or None
        supplier_phone = request.POST.get('supplier_phone', '').strip() or None
        supplier_address = request.POST.get('supplier_address', '').strip() or None
        document_number = request.POST.get('document_number', '').strip()
        tipo_pago = request.POST.get('tipo_pago', 'CREDITO').strip()
        
        product_ids = request.POST.getlist('product[]')
        quantities = request.POST.getlist('quantity[]')
        unit_costs = request.POST.getlist('unit_cost[]')
        
        errors = []
        if not supplier_select:
            errors.append("Debe seleccionar un proveedor o elegir '-- Nuevo Proveedor --'.")
            
        if not document_number:
            errors.append("El número de documento es obligatorio.")
            
        if supplier_select == 'new':
            if not supplier_name:
                errors.append("El nombre del proveedor es obligatorio para un nuevo proveedor.")
            if supplier_phone:
                phone_cleaned = re.sub(r'\s+|-', '', supplier_phone)
                if not re.match(r'^\d+$', phone_cleaned):
                    errors.append("El teléfono del proveedor solo debe contener números.")
        
        if not product_ids:
            errors.append("Debe agregar al menos un producto a la compra.")
            
        if len(product_ids) != len(quantities) or len(product_ids) != len(unit_costs):
            errors.append("Datos de productos incorrectos.")
            
        # Comprobar duplicado antes de guardar
        if supplier_select and supplier_select != 'new':
            if Purchase.objects.filter(supplier_id=supplier_select, document_number=document_number).exists():
                errors.append(f"El número de documento '{document_number}' ya está registrado para este proveedor.")
        elif supplier_select == 'new' and supplier_name:
            if Purchase.objects.filter(supplier__name=supplier_name, document_number=document_number).exists():
                errors.append(f"El número de documento '{document_number}' ya está registrado para el proveedor '{supplier_name}'.")

        items_to_save = []
        # Uso de decimal para evitar errores de redondeo
        subtotal = Decimal('0.00')
        
        for idx in range(len(product_ids)):
            p_id = product_ids[idx]
            qty_str = quantities[idx]
            cost_str = unit_costs[idx]
            
            if not p_id or not qty_str or not cost_str:
                errors.append(f"Fila {idx+1}: Complete todos los campos del producto.")
                continue
                
            try:
                qty = int(qty_str)
                cost = Decimal(cost_str)
            except (ValueError, TypeError):
                errors.append(f"Fila {idx+1}: Cantidad o costo no válidos.")
                continue
                
            if qty <= 0:
                errors.append(f"Fila {idx+1}: La cantidad debe ser mayor que cero.")
                continue
                
            if cost < 0:
                errors.append(f"Fila {idx+1}: El costo no puede ser negativo.")
                continue
                
            product = Product.objects.filter(id=p_id).first()
            if not product:
                errors.append(f"Fila {idx+1}: El producto seleccionado no existe.")
                continue
                
            line_subtotal = qty * cost
            subtotal += line_subtotal
            items_to_save.append({
                'product': product,
                'quantity': qty,
                'unit_cost': cost,
                'subtotal': line_subtotal
            })
            
        if errors:
            for err in errors:
                messages.error(request, err)
            return _render_form_with_errors(request, supplier_select, supplier_name, supplier_contact_name, supplier_email, supplier_phone, supplier_address, document_number, tipo_pago, product_ids, quantities, unit_costs)
            
        try:
            with transaction.atomic():
                if supplier_select and supplier_select != 'new':
                    supplier = Supplier.objects.get(id=supplier_select)
                    supplier.name = supplier_name
                    supplier.contact_name = supplier_contact_name
                    supplier.email = supplier_email
                    supplier.phone = supplier_phone
                    supplier.address = supplier_address
                    supplier.save()
                else:
                    supplier = Supplier.objects.create(
                        name=supplier_name,
                        contact_name=supplier_contact_name,
                        email=supplier_email,
                        phone=supplier_phone,
                        address=supplier_address
                    )
                
                # Verificar de nuevo que no exista
                if Purchase.objects.filter(supplier=supplier, document_number=document_number).exists():
                    raise Exception(f"El número de documento '{document_number}' ya está registrado para este proveedor.")
                
                tax = subtotal * Decimal('0.15')
                total = subtotal + tax
                
                purchase = Purchase.objects.create(
                    supplier=supplier,
                    document_number=document_number,
                    tipo_pago=tipo_pago,
                    subtotal=subtotal,
                    tax=tax,
                    total=total,
                    is_active=True
                )
                
                for item in items_to_save:
                    PurchaseDetail.objects.create(
                        purchase=purchase,
                        product=item['product'],
                        quantity=item['quantity'],
                        unit_cost=item['unit_cost'],
                        subtotal=item['subtotal']
                    )
                    # ACTUALIZAR STOCK: suma la cantidad comprada al stock del producto
                    Product.objects.filter(id=item['product'].id).update(stock=F('stock') + item['quantity'])
                    
            messages.success(request, f"Compra #{purchase.id} guardada con éxito. Stock actualizado.")
            return redirect('purchasing:purchase_list')
            
        except Exception as e:
            messages.error(request, f"Error al guardar la compra: {str(e)}")
            return _render_form_with_errors(request, supplier_select, supplier_name, supplier_contact_name, supplier_email, supplier_phone, supplier_address, document_number, tipo_pago, product_ids, quantities, unit_costs)
            
    else:
        return _render_form(request)

def _render_form(request, context_opts=None):
    if context_opts is None:
        context_opts = {}
    suppliers = Supplier.objects.filter(is_active=True)
    products = Product.objects.filter(is_active=True).select_related('brand')
    
    suppliers_json = []
    for s in suppliers:
        suppliers_json.append({
            'id': s.id,
            'name': s.name,
            'contact_name': s.contact_name or '',
            'email': s.email or '',
            'phone': s.phone or '',
            'address': s.address or '',
        })
        
    products_json = []
    for p in products:
        products_json.append({
            'id': p.id,
            'name': f"{p.name} ({p.brand.name})",
            'price': float(p.unit_price),
        })
        
    context = {
        'title': 'Crear Compra',
        'suppliers': suppliers,
        'products': products,
        'suppliers_json_str': json.dumps(suppliers_json),
        'products_json_str': json.dumps(products_json),
    }
    context.update(context_opts)
    return render(request, 'purchasing/purchase_form.html', context)

def _render_form_with_errors(request, supplier_select, supplier_name, supplier_contact_name, supplier_email, supplier_phone, supplier_address, document_number, tipo_pago, product_ids, quantities, unit_costs):
    posted_values = {
        'supplier_select': supplier_select,
        'supplier_name': supplier_name,
        'supplier_contact_name': supplier_contact_name,
        'supplier_email': supplier_email or '',
        'supplier_phone': supplier_phone or '',
        'supplier_address': supplier_address or '',
        'document_number': document_number,
        'tipo_pago': tipo_pago,
    }
    
    selected_items = []
    for idx in range(len(product_ids)):
        try:
            selected_items.append({
                'product_id': int(product_ids[idx]) if product_ids[idx] else '',
                'quantity': int(quantities[idx]) if quantities[idx] else 1,
                'unit_cost': float(unit_costs[idx]) if unit_costs[idx] else 0.00
            })
        except Exception:
            selected_items.append({
                'product_id': product_ids[idx],
                'quantity': quantities[idx],
                'unit_cost': unit_costs[idx]
            })
            
    return _render_form(request, {
        'posted_values_json': json.dumps(posted_values),
        'selected_items_json': json.dumps(selected_items),
    })

@login_required
@permission_required('purchasing.view_purchase')
def purchase_detail(request, pk):
    """Detalle de una compra con prefetch_related('details__product')."""
    purchase = get_object_or_404(
        Purchase.objects.select_related('supplier').prefetch_related('details__product'),
        pk=pk
    )
    return render(request, 'purchasing/purchase_detail.html', {'purchase': purchase})

@login_required
@permission_required('purchasing.delete_purchase')
def purchase_delete(request, pk):
    """Elimina una compra y todos sus detalles (CASCADE)."""
    purchase = get_object_or_404(Purchase, pk=pk)
    if request.method == 'POST':
        purchase_id = purchase.id
        purchase.delete()
        messages.success(request, f'¡Compra #{purchase_id} eliminada con éxito!')
        return redirect('purchasing:purchase_list')
    return render(request, 'purchasing/purchase_confirm_delete.html', {'object': purchase})

@login_required
@permission_required('purchasing.view_purchase')
def purchase_report(request):
    """Reporte de costo promedio por producto."""
    products = Product.objects.annotate(
        avg_cost=Avg('purchase_details__unit_cost')
    ).select_related('brand', 'group').order_by('name')
    
    overall_avg = PurchaseDetail.objects.aggregate(avg_cost=Avg('unit_cost'))['avg_cost'] or Decimal('0.00')
    
    context = {
        'products': products,
        'overall_avg': overall_avg,
    }
    return render(request, 'purchasing/purchase_report.html', context)


# === REPORTES GENÉRICOS HELPERS ===
from reportlab.pdfgen import canvas
from reportlab.lib import colors

class GlobalNumberedCanvas(canvas.Canvas):
    report_title = "Sistema de Ventas - Reporte de Compras"
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor("#7F7F7F"))
        self.setStrokeColor(colors.HexColor("#D9D9D9"))
        self.setLineWidth(0.5)
        self.line(36, 756, 576, 756)
        self.drawString(36, 762, self.report_title)
        self.line(36, 54, 576, 54)
        page_text = f"Página {self._pageNumber} de {page_count}"
        self.drawRightString(576, 42, page_text)
        self.drawString(36, 42, "Reporte generado automáticamente")
        self.restoreState()

def get_numbered_canvas_class(title):
    class DynamicNumberedCanvas(GlobalNumberedCanvas):
        report_title = title
    return DynamicNumberedCanvas


# === REPORTES DE COMPRAS (Excel / PDF) ===
@login_required
@permission_required('purchasing.view_purchase')
def purchase_report_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compras"
    
    ws.merge_cells("A1:H1")
    ws["A1"] = "REPORTE GENERAL DE COMPRAS (ADQUISICIONES)"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    headers = ["ID Compra", "Proveedor", "Nº Factura Proveedor", "Fecha Adquisición", "Subtotal (USD)", "IVA (15%)", "Total (USD)", "Estado"]
    ws.append([])
    ws.append(headers)
    
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    for col in range(1, 9):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    ws.row_dimensions[3].height = 25
    
    purchases = Purchase.objects.select_related('supplier').all().order_by('-purchase_date')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for p in purchases:
        row = [
            p.id,
            p.supplier.name,
            p.document_number,
            p.purchase_date.strftime("%d/%m/%Y %H:%M"),
            float(p.subtotal),
            float(p.tax),
            float(p.total),
            "Activo" if p.is_active else "Inactivo"
        ]
        ws.append(row)
        curr_row = ws.max_row
        ws.row_dimensions[curr_row].height = 20
        
        ws.cell(row=curr_row, column=1).alignment = center_align
        ws.cell(row=curr_row, column=2).alignment = left_align
        ws.cell(row=curr_row, column=3).alignment = center_align
        ws.cell(row=curr_row, column=4).alignment = center_align
        
        for col in [5, 6, 7]:
            val_cell = ws.cell(row=curr_row, column=col)
            val_cell.alignment = right_align
            val_cell.number_format = '$#,##0.00'
            
        ws.cell(row=curr_row, column=8).alignment = center_align
        
        for col in range(1, 9):
            c = ws.cell(row=curr_row, column=col)
            c.font = Font(name="Calibri", size=11)
            c.border = thin_border
            
    ws.auto_filter.ref = f"A3:H{ws.max_row}"
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            if cell.column in [5, 6, 7] and isinstance(cell.value, (int, float)):
                val_str = f"${cell.value:,.2f}"
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Reporte_Compras.xlsx"'
    wb.save(response)
    return response


@login_required
@permission_required('purchasing.view_purchase')
def purchase_report_pdf(request):
    import io
    import datetime
    from django.http import FileResponse
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=54,
        bottomMargin=72
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=22,
        textColor=colors.HexColor("#1F4E78"),
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor("#595959"),
        spaceAfter=25
    )
    cell_header_style = ParagraphStyle(
        'CellHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.white,
        alignment=1
    )
    cell_text_left = ParagraphStyle(
        'CellTextLeft',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor("#262626")
    )
    cell_text_center = ParagraphStyle(
        'CellTextCenter',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor("#262626"),
        alignment=1
    )
    cell_text_right = ParagraphStyle(
        'CellTextRight',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor("#262626"),
        alignment=2
    )
    
    now_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    story.append(Paragraph("REPORTE GENERAL DE COMPRAS (ADQUISICIONES)", title_style))
    story.append(Paragraph(f"Fecha de Generación: {now_str} | Historial de compras registradas.", subtitle_style))
    
    purchases = Purchase.objects.select_related('supplier').all().order_by('-purchase_date')
    
    data = [
        [
            Paragraph("ID Compra", cell_header_style),
            Paragraph("Proveedor", cell_header_style),
            Paragraph("Nº Factura Prov.", cell_header_style),
            Paragraph("Fecha", cell_header_style),
            Paragraph("Subtotal", cell_header_style),
            Paragraph("IVA (15%)", cell_header_style),
            Paragraph("Total", cell_header_style)
        ]
    ]
    
    for p in purchases:
        data.append([
            Paragraph(f"#{p.id}", cell_text_center),
            Paragraph(p.supplier.name, cell_text_left),
            Paragraph(p.document_number, cell_text_center),
            Paragraph(p.purchase_date.strftime("%d/%m/%Y %H:%M"), cell_text_center),
            Paragraph(f"${p.subtotal:,.2f}", cell_text_right),
            Paragraph(f"${p.tax:,.2f}", cell_text_right),
            Paragraph(f"${p.total:,.2f}", cell_text_right)
        ])
        
    col_widths = [55, 140, 95, 90, 50, 50, 60]
    
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2F5597")),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#F2F4F7"), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))
    
    story.append(t)
    canvas_class = get_numbered_canvas_class("Sistema de Ventas - Reporte de Compras")
    doc.build(story, canvasmaker=canvas_class)
    
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename="Reporte_Compras.pdf")


def generate_purchase_pdf_data(purchase):
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=54,
        bottomMargin=54
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'PurchaseTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor("#2F5597"),
        spaceAfter=5
    )
    normal_style = ParagraphStyle(
        'PurchaseNormal',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor("#262626"),
        spaceAfter=3
    )
    bold_style = ParagraphStyle(
        'PurchaseBold',
        parent=normal_style,
        fontName='Helvetica-Bold'
    )
    
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.white,
        alignment=1
    )
    
    cell_left = ParagraphStyle(
        'CellLeft',
        parent=normal_style,
        fontSize=9
    )
    cell_center = ParagraphStyle(
        'CellCenter',
        parent=normal_style,
        fontSize=9,
        alignment=1
    )
    cell_right = ParagraphStyle(
        'CellRight',
        parent=normal_style,
        fontSize=9,
        alignment=2
    )

    story.append(Paragraph("DETALLE DE COMPRA / ADQUISICIÓN", title_style))
    story.append(Paragraph(f"Compra Nº: {purchase.id}", bold_style))
    story.append(Paragraph(f"Nº Factura Proveedor: {purchase.document_number}", bold_style))
    story.append(Paragraph(f"Fecha: {purchase.purchase_date.strftime('%d/%m/%Y %H:%M')}", normal_style))
    story.append(Spacer(1, 15))
    
    info_data = [
        [
            Paragraph("<b>COMPRADOR (EMPRESA)</b><br/>Sistema de Ventas<br/>adquisiciones@sistema.com", normal_style),
            Paragraph(f"<b>PROVEEDOR</b><br/><b>Nombre:</b> {purchase.supplier.name}<br/><b>Contacto:</b> {purchase.supplier.contact_name or '-'}<br/><b>Email:</b> {purchase.supplier.email or '-'}<br/><b>Tlf:</b> {purchase.supplier.phone or '-'}", normal_style)
        ]
    ]
    info_table = Table(info_data, colWidths=[250, 250])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('PADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 20))
    
    table_data = [
        [
            Paragraph("Producto", header_style),
            Paragraph("Cantidad", header_style),
            Paragraph("Costo Unitario", header_style),
            Paragraph("Subtotal", header_style)
        ]
    ]
    
    for detail in purchase.details.all():
        table_data.append([
            Paragraph(detail.product.name, cell_left),
            Paragraph(str(detail.quantity), cell_center),
            Paragraph(f"${detail.unit_cost:,.2f}", cell_right),
            Paragraph(f"${detail.subtotal:,.2f}", cell_right)
        ])
        
    items_table = Table(table_data, colWidths=[260, 60, 90, 90])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#2F5597")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#D9D9D9")),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(items_table)
    story.append(Spacer(1, 15))
    
    totals_data = [
        [Paragraph("", cell_left), Paragraph("Subtotal:", cell_right), Paragraph(f"${purchase.subtotal:,.2f}", cell_right)],
        [Paragraph("", cell_left), Paragraph("IVA (15%):", cell_right), Paragraph(f"${purchase.tax:,.2f}", cell_right)],
        [Paragraph("", cell_left), Paragraph("TOTAL:", ParagraphStyle('TBold', parent=cell_right, fontName='Helvetica-Bold')), Paragraph(f"${purchase.total:,.2f}", ParagraphStyle('TBold2', parent=cell_right, fontName='Helvetica-Bold'))]
    ]
    totals_table = Table(totals_data, colWidths=[320, 90, 90])
    totals_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LINEABOVE', (1,2), (2,2), 1, colors.HexColor("#2F5597")),
    ]))
    story.append(totals_table)
    
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


@login_required
@permission_required('purchasing.view_purchase')
def purchase_pdf(request, pk):
    from django.http import HttpResponse
    purchase = get_object_or_404(Purchase, pk=pk)
    pdf_data = generate_purchase_pdf_data(purchase)
    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Compra_{purchase.id}.pdf"'
    return response


