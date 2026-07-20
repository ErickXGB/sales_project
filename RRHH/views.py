from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.db import transaction
from django.shortcuts import redirect
from django.http import JsonResponse
from django.db.models import Sum, Count

from shared.mixins import PermissionRequiredMixin
from shared.decorators import permission_required

from decimal import Decimal

from .models import Sobretiempo, Empleado, TipoSobretiempo, Prestamo, PrestamoDetalle, TipoPrestamo
from .forms import SobretiempoForm, SobretiempoDetalleFormSet, PrestamoForm, PrestamoDetalleFormSet


class SobretiempoListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'RRHH.view_sobretiempo'
    model = Sobretiempo
    template_name = 'rrhh/sobretiempo_list.html'
    context_object_name = 'sobretiempos'


class SobretiempoDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    permission_required = 'RRHH.detail_sobretiempo'
    model = Sobretiempo
    template_name = 'rrhh/sobretiempo_detail.html'
    context_object_name = 'sobretiempo'


class SobretiempoCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'RRHH.add_sobretiempo'
    model = Sobretiempo
    form_class = SobretiempoForm
    template_name = 'rrhh/sobretiempo_form.html'
    success_url = reverse_lazy('rrhh:sobretiempo_list')

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        # Convertir Decimal a float para evitar errores en JavaScript
        data['empleados_json'] = [{'id': e.id, 'sueldo': float(e.sueldo)} for e in Empleado.objects.all()]
        data['tipos_json'] = [{'id': t.id, 'factor': float(t.factor)} for t in TipoSobretiempo.objects.all()]
        
        if 'detalles' in kwargs:
            data['detalles'] = kwargs['detalles']
        elif self.request.POST:
            instance = getattr(self, 'object', None)
            data['detalles'] = SobretiempoDetalleFormSet(self.request.POST, instance=instance)
        else:
            instance = getattr(self, 'object', None)
            data['detalles'] = SobretiempoDetalleFormSet(instance=instance)
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        detalles = context['detalles']
        
        # Validar que cabecera y detalles se guardan juntos
        with transaction.atomic():
            self.object = form.save()
            detalles.instance = self.object
            if detalles.is_valid():
                detalles.save()
                
                # Forzar recalcular por seguridad
                self.object.calcular_total_maestro()
                messages.success(self.request, f"Sobretiempo registrado con éxito.")
                return super().form_valid(form)
            else:
                # Si los detalles son inválidos, volver a intentar pasando detalles en kwargs
                return self.render_to_response(self.get_context_data(form=form, detalles=detalles))


class SobretiempoUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'RRHH.change_sobretiempo'
    model = Sobretiempo
    form_class = SobretiempoForm
    template_name = 'rrhh/sobretiempo_form.html'
    success_url = reverse_lazy('rrhh:sobretiempo_list')

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        data['empleados_json'] = [{'id': e.id, 'sueldo': float(e.sueldo)} for e in Empleado.objects.all()]
        data['tipos_json'] = [{'id': t.id, 'factor': float(t.factor)} for t in TipoSobretiempo.objects.all()]
        
        if 'detalles' in kwargs:
            data['detalles'] = kwargs['detalles']
        elif self.request.POST:
            data['detalles'] = SobretiempoDetalleFormSet(self.request.POST, instance=self.object)
        else:
            data['detalles'] = SobretiempoDetalleFormSet(instance=self.object)
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        detalles = context['detalles']
        
        with transaction.atomic():
            self.object = form.save()
            detalles.instance = self.object
            if detalles.is_valid():
                detalles.save()
                
                # Recalcular
                self.object.calcular_total_maestro()
                messages.success(self.request, f"Registro de sobretiempo actualizado con éxito.")
                return super().form_valid(form)
            else:
                return self.render_to_response(self.get_context_data(form=form, detalles=detalles))


class SobretiempoDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = 'RRHH.delete_sobretiempo'
    model = Sobretiempo
    template_name = 'rrhh/sobretiempo_confirm_delete.html'
    success_url = reverse_lazy('rrhh:sobretiempo_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Registro de sobretiempo eliminado correctamente.")
        return super().delete(request, *args, **kwargs)


class SobretiempoResumenView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'RRHH.view_sobretiempo'
    model = Sobretiempo
    template_name = 'rrhh/sobretiempo_resumen.html'
    context_object_name = 'sobretiempos'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Sumar todos los montos calculados del sistema
        totales = Sobretiempo.objects.aggregate(
            total_general=Sum('total_calculado'),
            total_registros=Count('id')
        )
        context['total_general'] = totales['total_general'] or 0
        context['total_registros'] = totales['total_registros'] or 0

        # Resumen de totales agrupados por empleado
        resumen_empleados = (
            Sobretiempo.objects.values('empleado__nombres')
            .annotate(
                total_horas_extras=Sum('detalles__numero_horas'),
                monto_total=Sum('total_calculado'),
                cantidad_registros=Count('id', distinct=True)
            )
            .order_by('-monto_total')
        )
        context['resumen_empleados'] = resumen_empleados
        return context


@permission_required('RRHH.download_sobretiempo_pdf')
def export_sobretiempo_list_pdf(request):
    import io
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from django.http import HttpResponse
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=54,
        bottomMargin=54
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor("#1F4E78"),
        spaceAfter=15
    )
    
    story.append(Paragraph("Reporte General de Sobretiempos", title_style))
    story.append(Spacer(1, 10))
    
    headers = ["Empleado", "Fecha Registro", "Horas del Mes", "Sueldo Mensual", "Total Calculado"]
    data = [headers]
    
    sobretiempos = Sobretiempo.objects.all().order_by('fecha_registro')
    for s in sobretiempos:
        data.append([
            s.empleado.nombres,
            s.fecha_registro.strftime('%d/%m/%Y'),
            str(s.total_horas),
            f"${s.sueldo_mensual:.2f}",
            f"${s.total_calculado:.2f}"
        ])
        
    table = Table(data, colWidths=[180, 90, 80, 90, 100])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2F5597")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#F2F2F2")),
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_general_sobretiempos.pdf"'
    return response


@permission_required('RRHH.download_sobretiempo_excel')
def export_sobretiempo_list_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sobretiempos"
    
    ws.merge_cells("A1:E1")
    ws["A1"] = "REPORTE GENERAL DE SOBRETIEMPOS"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    headers = ["Empleado", "Fecha de Registro", "Horas del Mes (Divisor)", "Sueldo Mensual (USD)", "Total Calculado (USD)"]
    ws.append([])
    ws.append(headers)
    
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, 6):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    ws.row_dimensions[3].height = 25
    
    sobretiempos = Sobretiempo.objects.all().order_by('fecha_registro')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for s in sobretiempos:
        ws.append([
            s.empleado.nombres,
            s.fecha_registro.strftime('%d/%m/%Y'),
            s.total_horas,
            float(s.sueldo_mensual),
            float(s.total_calculado)
        ])
        
    for r in range(4, ws.max_row + 1):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            if c in [3, 4, 5]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            if c == 4 or c == 5:
                cell.number_format = "$#,##0.00"
                
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="reporte_general_sobretiempos.xlsx"'
    wb.save(response)
    return response


@permission_required('RRHH.download_sobretiempo_pdf')
def export_sobretiempo_detail_pdf(request, pk):
    import io
    from django.shortcuts import get_object_or_404
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from django.http import HttpResponse
    
    sobretiempo = get_object_or_404(Sobretiempo, pk=pk)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=54,
        bottomMargin=54
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor("#1F4E78"),
        spaceAfter=15
    )
    
    story.append(Paragraph(f"Detalle de Sobretiempo #{sobretiempo.id}", title_style))
    story.append(Spacer(1, 10))
    
    cabecera_data = [
        ["Empleado:", sobretiempo.empleado.nombres, "Fecha Registro:", sobretiempo.fecha_registro.strftime('%d/%m/%Y')],
        ["Sueldo Mensual:", f"${sobretiempo.sueldo_mensual:.2f}", "Horas Mensuales (Divisor):", str(sobretiempo.total_horas)]
    ]
    cabecera_table = Table(cabecera_data, colWidths=[120, 150, 150, 120])
    cabecera_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    story.append(cabecera_table)
    story.append(Spacer(1, 20))
    story.append(Paragraph("Desglose de Horas Extras Realizadas", ParagraphStyle('Sub', parent=styles['Heading2'], fontSize=12, spaceAfter=8)))
    
    headers = ["Tipo de Sobretiempo", "Factor", "Número de Horas", "Valor Calculado"]
    data = [headers]
    
    for d in sobretiempo.detalles.all():
        data.append([
            d.tipo_sobretiempo.descripcion,
            f"{d.tipo_sobretiempo.factor:.2f}",
            f"{d.numero_horas:.2f} hrs",
            f"${d.valor_calculado:.2f}"
        ])
        
    data.append(["", "", "Total Acumulado:", f"${sobretiempo.total_calculado:.2f}"])
    
    table = Table(data, colWidths=[200, 90, 120, 130])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2F5597")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
        ('LINEBELOW', (0, -1), (-1, -1), 1, colors.HexColor("#1F4E78")),
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTNAME', (2, -1), (3, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor("#F2F2F2")),
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="sobretiempo_{sobretiempo.id}_{sobretiempo.empleado.nombres.replace(" ", "_")}.pdf"'
    return response


@permission_required('RRHH.download_sobretiempo_excel')
def export_sobretiempo_detail_excel(request, pk):
    import openpyxl
    from django.shortcuts import get_object_or_404
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    
    sobretiempo = get_object_or_404(Sobretiempo, pk=pk)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Sobretiempo #{sobretiempo.id}"
    
    ws.merge_cells("A1:D1")
    ws["A1"] = f"DETALLE DE SOBRETIEMPO #{sobretiempo.id}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    
    ws.append([])
    
    ws.append(["Empleado:", sobretiempo.empleado.nombres, "Fecha Registro:", sobretiempo.fecha_registro.strftime('%d/%m/%Y')])
    ws.append(["Sueldo Mensual:", float(sobretiempo.sueldo_mensual), "Horas Mensuales:", sobretiempo.total_horas])
    
    bold_font = Font(name="Calibri", size=11, bold=True)
    ws["A3"].font = bold_font
    ws["C3"].font = bold_font
    ws["A4"].font = bold_font
    ws["C4"].font = bold_font
    ws["B4"].number_format = "$#,##0.00"
    
    ws.append([])
    ws.append([])
    
    headers = ["Tipo de Sobretiempo", "Factor de Recargo", "Número de Horas", "Valor Calculado (USD)"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    for col in range(1, 5):
        cell = ws.cell(row=7, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 25
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for d in sobretiempo.detalles.all():
        ws.append([
            d.tipo_sobretiempo.descripcion,
            float(d.tipo_sobretiempo.factor),
            float(d.numero_horas),
            float(d.valor_calculado)
        ])
        
    ws.append(["", "", "Total Acumulado:", float(sobretiempo.total_calculado)])
    
    last_row = ws.max_row
    ws.cell(row=last_row, column=3).font = bold_font
    ws.cell(row=last_row, column=4).font = bold_font
    ws.cell(row=last_row, column=4).number_format = "$#,##0.00"
    
    for r in range(8, last_row):
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            if c in [2, 3, 4]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            if c == 4:
                cell.number_format = "$#,##0.00"
                
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="sobretiempo_{sobretiempo.id}.xlsx"'
    wb.save(response)
    return response


# ==========================================
# VISTAS DE GESTIÓN DE PRÉSTAMOS
# ==========================================

class PrestamoListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'RRHH.view_prestamo'
    model = Prestamo
    template_name = 'rrhh/prestamo_list.html'
    context_object_name = 'prestamos'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        totales = Prestamo.objects.aggregate(
            total_monto=Sum('monto'),
            total_pagar=Sum('monto_pagar'),
            total_saldo=Sum('saldo'),
            total_registros=Count('id')
        )
        context['total_monto'] = totales['total_monto'] or 0
        context['total_pagar'] = totales['total_pagar'] or 0
        context['total_saldo'] = totales['total_saldo'] or 0
        context['total_registros'] = totales['total_registros'] or 0
        return context


class PrestamoDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    permission_required = 'RRHH.detail_prestamo'
    model = Prestamo
    template_name = 'rrhh/prestamo_detail.html'
    context_object_name = 'prestamo'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        proxima = self.object.detalles.filter(saldo_cuota__gt=Decimal('0.00')).order_by('numero_cuota').first()
        context['proxima_cuota_id'] = proxima.id if proxima else None
        return context


@permission_required('RRHH.change_prestamo')
def pagar_cuota_prestamo(request, cuota_id):
    from django.shortcuts import get_object_or_404
    cuota = get_object_or_404(PrestamoDetalle, pk=cuota_id)
    
    # Validar que no se puedan pagar cuotas futuras si hay cuotas anteriores pendientes
    cuota_ant_pendiente = cuota.prestamo.detalles.filter(
        numero_cuota__lt=cuota.numero_cuota,
        saldo_cuota__gt=Decimal('0.00')
    ).order_by('numero_cuota').first()

    if cuota_ant_pendiente:
        messages.error(
            request,
            f"No puede pagar la cuota #{cuota.numero_cuota} sin haber pagado previamente la cuota #{cuota_ant_pendiente.numero_cuota}."
        )
        return redirect('rrhh:prestamo_detail', pk=cuota.prestamo.id)

    cuota.saldo_cuota = Decimal('0.00')
    cuota.save()
    messages.success(request, f"La cuota #{cuota.numero_cuota} del préstamo #{cuota.prestamo.id} ha sido marcada como pagada.")
    return redirect('rrhh:prestamo_detail', pk=cuota.prestamo.id)



class PrestamoCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'RRHH.add_prestamo'
    model = Prestamo
    form_class = PrestamoForm
    template_name = 'rrhh/prestamo_form.html'
    success_url = reverse_lazy('rrhh:prestamo_list')

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        data['tipos_prestamo_json'] = [
            {'id': t.id, 'descripcion': t.descripcion, 'tasa_interes': t.tasa_interes}
            for t in TipoPrestamo.objects.all()
        ]
        if 'detalles' in kwargs:
            data['detalles'] = kwargs['detalles']
        elif self.request.POST:
            instance = getattr(self, 'object', None)
            data['detalles'] = PrestamoDetalleFormSet(self.request.POST, instance=instance)
        else:
            instance = getattr(self, 'object', None)
            data['detalles'] = PrestamoDetalleFormSet(instance=instance)
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        detalles = context['detalles']

        with transaction.atomic():
            self.object = form.save()
            if detalles.is_valid():
                if detalles.has_changed():
                    detalles.instance = self.object
                    detalles.save()
                self.object.actualizar_saldo_y_estado()
                messages.success(self.request, f"Préstamo registrado con éxito.")
                return redirect(self.success_url)
            else:
                return self.render_to_response(self.get_context_data(form=form, detalles=detalles))

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))


class PrestamoUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'RRHH.change_prestamo'
    model = Prestamo
    form_class = PrestamoForm
    template_name = 'rrhh/prestamo_form.html'
    success_url = reverse_lazy('rrhh:prestamo_list')

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        data['tipos_prestamo_json'] = [
            {'id': t.id, 'descripcion': t.descripcion, 'tasa_interes': t.tasa_interes}
            for t in TipoPrestamo.objects.all()
        ]
        if 'detalles' in kwargs:
            data['detalles'] = kwargs['detalles']
        elif self.request.POST:
            data['detalles'] = PrestamoDetalleFormSet(self.request.POST, instance=self.object)
        else:
            data['detalles'] = PrestamoDetalleFormSet(instance=self.object)
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        detalles = context['detalles']

        with transaction.atomic():
            self.object = form.save()
            if detalles.is_valid():
                detalles.instance = self.object
                detalles.save()

                self.object.actualizar_saldo_y_estado()
                messages.success(self.request, f"Préstamo actualizado con éxito.")
                return redirect(self.success_url)
            else:
                return self.render_to_response(self.get_context_data(form=form, detalles=detalles))



class PrestamoDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = 'RRHH.delete_prestamo'
    model = Prestamo
    template_name = 'rrhh/prestamo_confirm_delete.html'
    success_url = reverse_lazy('rrhh:prestamo_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Préstamo eliminado correctamente.")
        return super().delete(request, *args, **kwargs)






class PrestamoResumenView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'RRHH.view_prestamo'
    model = Prestamo
    template_name = 'rrhh/prestamo_resumen.html'
    context_object_name = 'prestamos'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        totales = Prestamo.objects.aggregate(
            total_solicitado=Sum('monto'),
            total_interes=Sum('interes'),
            total_pagar=Sum('monto_pagar'),
            total_saldo=Sum('saldo'),
            total_registros=Count('id')
        )
        context['totales'] = totales
        resumen_empleados = (
            Prestamo.objects.values('empleado__nombres')
            .annotate(
                total_prestamos=Count('id'),
                monto_total=Sum('monto'),
                monto_pagar_total=Sum('monto_pagar'),
                saldo_pendiente=Sum('saldo')
            )
            .order_by('-monto_total')
        )
        context['resumen_empleados'] = resumen_empleados
        return context


@permission_required('RRHH.download_prestamo_pdf')
def export_prestamo_list_pdf(request):
    import io
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from django.http import HttpResponse
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=54,
        bottomMargin=54
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor("#1F4E78"),
        spaceAfter=15
    )
    
    story.append(Paragraph("Reporte General de Préstamos", title_style))
    story.append(Spacer(1, 10))
    
    headers = ["ID", "Empleado", "Tipo Préstamo", "Fecha", "Monto", "Interés", "Total Pagar", "Saldo", "Estado"]
    data = [headers]
    
    prestamos = Prestamo.objects.all().order_by('-fecha_prestamo')
    for p in prestamos:
        data.append([
            str(p.id),
            p.empleado.nombres,
            p.tipo_prestamo.descripcion,
            p.fecha_prestamo.strftime('%d/%m/%Y'),
            f"${p.monto:.2f}",
            f"${p.interes:.2f}",
            f"${p.monto_pagar:.2f}",
            f"${p.saldo:.2f}",
            p.get_estado_display()
        ])
        
    table = Table(data, colWidths=[30, 110, 85, 65, 60, 55, 65, 60, 50])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#F9F9F9")),
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_general_prestamos.pdf"'
    return response


@permission_required('RRHH.download_prestamo_excel')
def export_prestamo_list_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Préstamos"
    
    ws.merge_cells("A1:I1")
    ws["A1"] = "REPORTE GENERAL DE PRÉSTAMOS"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    headers = ["ID", "Empleado", "Tipo de Préstamo", "Fecha Préstamo", "Monto ($)", "Interés ($)", "Monto Total ($)", "Saldo ($)", "Estado"]
    ws.append([])
    ws.append(headers)
    
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    for col in range(1, 10):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 25
    
    prestamos = Prestamo.objects.all().order_by('-fecha_prestamo')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for p in prestamos:
        ws.append([
            p.id,
            p.empleado.nombres,
            p.tipo_prestamo.descripcion,
            p.fecha_prestamo.strftime('%d/%m/%Y'),
            float(p.monto),
            float(p.interes),
            float(p.monto_pagar),
            float(p.saldo),
            p.get_estado_display()
        ])
        
    for r in range(4, ws.max_row + 1):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            if c in [5, 6, 7, 8]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "$#,##0.00"
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="reporte_general_prestamos.xlsx"'
    wb.save(response)
    return response


@permission_required('RRHH.download_prestamo_pdf')
def export_prestamo_detail_pdf(request, pk):
    import io
    from django.shortcuts import get_object_or_404
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from django.http import HttpResponse
    
    prestamo = get_object_or_404(Prestamo, pk=pk)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=54,
        bottomMargin=54
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor("#1F4E78"),
        spaceAfter=15
    )
    
    story.append(Paragraph(f"Tabla de Amortización - Préstamo #{prestamo.id}", title_style))
    story.append(Spacer(1, 10))
    
    cabecera_data = [
        ["Empleado:", prestamo.empleado.nombres, "Fecha Préstamo:", prestamo.fecha_prestamo.strftime('%d/%m/%Y')],
        ["Tipo Préstamo:", prestamo.tipo_prestamo.descripcion, "Tasa Interés:", f"{prestamo.tipo_prestamo.tasa_interes}%"],
        ["Monto Solicitado:", f"${prestamo.monto:.2f}", "Interés Generado:", f"${prestamo.interes:.2f}"],
        ["Monto Total Pagar:", f"${prestamo.monto_pagar:.2f}", "Saldo Pendiente:", f"${prestamo.saldo:.2f}"],
        ["Número Cuotas:", str(prestamo.numero_cuotas), "Estado:", prestamo.get_estado_display()]
    ]
    cabecera_table = Table(cabecera_data, colWidths=[120, 150, 140, 130])
    cabecera_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    story.append(cabecera_table)
    story.append(Spacer(1, 15))
    story.append(Paragraph("Detalle de Cuotas del Préstamo", ParagraphStyle('Sub', parent=styles['Heading2'], fontSize=12, spaceAfter=8)))
    
    headers = ["N° Cuota", "Fecha Vencimiento", "Valor Cuota", "Saldo Cuota", "Estado"]
    data = [headers]
    
    for d in prestamo.detalles.all():
        est = "PAGADA" if d.saldo_cuota <= 0 else "PENDIENTE"
        data.append([
            f"Cuota #{d.numero_cuota}",
            d.fecha_vencimiento.strftime('%d/%m/%Y'),
            f"${d.valor_cuota:.2f}",
            f"${d.saldo_cuota:.2f}",
            est
        ])
        
    table = Table(data, colWidths=[100, 150, 120, 120, 90])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#F9F9F9")),
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="prestamo_{prestamo.id}_{prestamo.empleado.nombres.replace(" ", "_")}.pdf"'
    return response

